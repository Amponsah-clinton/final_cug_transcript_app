# metascholar
# Meta
import csv
import hashlib
import hmac
import json
import openpyxl
import secrets
import unicodedata
import urllib.parse
import uuid
from datetime import datetime, timedelta
from io import BytesIO, TextIOWrapper
import requests
from reportlab.graphics.barcode import code128
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from django.conf import settings
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth import login, logout, views as auth_views
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib.auth.tokens import default_token_generator
from django.db.models import OuterRef, Prefetch, Q, Subquery
from django.http import (
    FileResponse,
    HttpResponse,
    HttpResponseNotFound,
    JsonResponse,
)
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.utils.http import urlsafe_base64_decode
import http.client
from django.urls import reverse
from django.utils.text import slugify
from django.views.decorators.csrf import csrf_exempt
from django.core.mail import send_mail, EmailMultiAlternatives
from .forms import (
    BulkStudentUploadForm,
    CustomSignupForm,
    FeeClearanceForm,
    PasswordResetConfirmForm,
    PasswordResetRequestForm,
    PaymentForm,
    StaffSignupForm,
    StudentForm,
    StudentUploadForm,
    TranscriptApprovalForm,
    TranscriptForm,
    TranscriptRequestForm,
    TranscriptStatusForm,
)
from .models import (
    Department,
    FacultyRegistrar,
    FeeClearance,
    PasswordResetCode,
    Payment,
    Program,
    StaffProfile,
    Student,
    StudentProfile,
    Transcript,
    TranscriptApproval,
    TranscriptRequest,
    TranscriptStatus,
    TranscriptVerification,
)
from .utils import generate_unofficial_transcript_pdf
from .utils import generate_official_transcript_pdf, load_signature_image
from .forms import RegistrarUploadForm

from django.views.decorators.http import require_GET


import uuid
from io import BytesIO
from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.http import require_GET
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.core.files.base import ContentFile
from django.db.models import Q
from PyPDF2 import PdfReader, PdfWriter

from .models import (
    Student, TranscriptRequest, Transcript, Payment, TranscriptVerification,
    TranscriptApproval, TranscriptStatus, TranscriptType, StaffProfile
)
from .forms import RegistrarUploadForm
from .utils import generate_official_transcript_pdf, generate_unofficial_transcript_pdf


@require_GET
def ajax_student_search(request):
    q = request.GET.get('q', '').strip()
    results = []
    if q:
        matches = Student.objects.filter(
            Q(index_number__icontains=q) | Q(name__icontains=q)
        ).order_by('index_number')[:20]
        for s in matches:
            results.append({'id': s.id, 'name': s.name, 'index_number': s.index_number})
    return JsonResponse({'results': results})


@login_required
def registrar_manual_upload(request):
    """Registrar manual upload not tied to an existing TranscriptRequest."""
    staff = StaffProfile.objects.filter(user=request.user, role='registrar').first()
    if not staff:
        messages.error(request, "Only Registrar staff can perform manual uploads.")
        return redirect('staff_dashboard')

    if request.method == 'POST':
        form = RegistrarUploadForm(request.POST, request.FILES)
        if form.is_valid():
            student_index = form.cleaned_data.get('student_index')
            student = Student.objects.filter(index_number__iexact=student_index).first()

            if not student:
                messages.error(request, "No student found with that index number.")
                return render(request, 'registrar_upload.html', {'form': form})

            # Always create a NEW TranscriptRequest for manual uploads so each manual upload has its own REF
            t_type = form.cleaned_data.get('transcript_type')
            tr = TranscriptRequest.objects.create(
                student=student,
                transcript_type=t_type,
                reference_code=f"REF-{uuid.uuid4().hex[:8].upper()}",
                payment_made=True if form.cleaned_data.get('amount_received') else False,
                amount=TranscriptType.objects.filter(type=t_type).first().price if TranscriptType.objects.filter(type=t_type).exists() else 0
            )

            # Attach uploaded file (original)
            uploaded_pdf = form.cleaned_data.get('transcript_file')
            transcript_obj, _ = Transcript.objects.get_or_create(transcript_request=tr)
            if uploaded_pdf:
                try:
                    transcript_obj.uploaded_file.save(
                        uploaded_pdf.name or f"uploaded_{uuid.uuid4().hex[:6]}.pdf",
                        uploaded_pdf,
                        save=False
                    )
                except Exception:
                    pass

            # Generate transcript PDF (uses tr for content so REF appears in PDF)
            base_url = request.build_absolute_uri('/')
            faculty_registrar = form.cleaned_data.get('faculty_registrar')
            if t_type == 'unofficial':
                transcript_obj, pdf_bytes = generate_unofficial_transcript_pdf(tr, faculty_registrar, base_url=base_url)
            else:
                transcript_obj, pdf_bytes = generate_official_transcript_pdf(tr, include_registrar=True, include_vc=True, base_url=base_url)

            # Merge uploaded PDF (if provided) after generated pages and save as canonical transcript.file
            if pdf_bytes:
                try:
                    merged_writer = PdfWriter()
                    gen_reader = PdfReader(BytesIO(pdf_bytes))
                    for page in gen_reader.pages:
                        merged_writer.add_page(page)

                    if uploaded_pdf:
                        if hasattr(uploaded_pdf, 'temporary_file_path'):
                            with open(uploaded_pdf.temporary_file_path(), 'rb') as uf:
                                up_reader = PdfReader(uf)
                                for page in up_reader.pages:
                                    merged_writer.add_page(page)
                        else:
                            uploaded_pdf.seek(0)
                            up_reader = PdfReader(uploaded_pdf)
                            for page in up_reader.pages:
                                merged_writer.add_page(page)

                    merged_output = BytesIO()
                    merged_writer.write(merged_output)
                    merged_output.seek(0)

                    merged_filename = f"{tr.reference_code}_{t_type}_combined.pdf"
                    # remove previous files if present
                    try:
                        if transcript_obj.file:
                            transcript_obj.file.delete(save=False)
                    except Exception:
                        pass
                    transcript_obj.file.save(merged_filename, ContentFile(merged_output.read()), save=True)
                except Exception as e:
                    print("PDF merge/save failed:", e)

            # Record payment breakdown
            payment, _ = Payment.objects.get_or_create(transcript_request=tr)
            payment.amount_received = form.cleaned_data.get('amount_received')
            payment.amount_accounts_office = form.cleaned_data.get('amount_accounts_office')
            payment.amount_superadmin = form.cleaned_data.get('amount_superadmin')
            payment.amount_registrar = form.cleaned_data.get('amount_registrar')
            payment.officer_name = request.user.get_full_name() or request.user.username
            payment.cleared = True
            payment.notes = (payment.notes or '') + f"\nManual upload by registrar."
            payment.save()

            # Verification and approval
            if pdf_bytes and transcript_obj:
                ver_obj, created = TranscriptVerification.objects.update_or_create(
                    transcript=transcript_obj,
                    defaults={
                        'barcode': f"VER-{transcript_obj.transcript_id.hex[:12].upper()}",
                        'verified': True if t_type == 'official' else False,
                        'date_verified': timezone.now() if t_type == 'official' else None,
                    }
                )

                if t_type == 'official':
                    TranscriptApproval.objects.update_or_create(
                        transcript=transcript_obj,
                        defaults={
                            'approved': True,
                            'approved_by': request.user.get_full_name() or request.user.username,
                            'date_approved': timezone.now(),
                        }
                    )
                    TranscriptStatus.objects.create(
                        transcript_request=tr,
                        stage='approved',
                        remarks="Manual official upload by Registrar.",
                        updated_by=request.user.get_full_name() or request.user.username
                    )
                else:
                    TranscriptStatus.objects.create(
                        transcript_request=tr,
                        stage='registrar',
                        remarks=f"Manual unofficial upload by Registrar. Faculty: {getattr(faculty_registrar, 'name', '')}",
                        updated_by=request.user.get_full_name() or request.user.username
                    )

            messages.success(request, 'Manual upload processed and merged successfully.')
            return redirect('staff_dashboard')
        else:
            messages.error(request, 'Please correct the form errors.')
    else:
        form = RegistrarUploadForm()

    return render(request, 'registrar_upload.html', {'form': form, 'manual': True})


@login_required
def exams_office_manual_upload(request):
    """Exams office manual upload that forwards to registrar for approval."""
    staff = StaffProfile.objects.filter(user=request.user, role='exams_office').first()
    if not staff:
        messages.error(request, "Only Exams Office staff can perform manual uploads.")
        return redirect('staff_dashboard')

    if request.method == 'POST':
        from .forms import ExamsOfficeUploadForm
        form = ExamsOfficeUploadForm(request.POST, request.FILES)
        if form.is_valid():
            student_index = form.cleaned_data.get('student_index')
            student = Student.objects.filter(index_number__iexact=student_index).first()

            if not student:
                messages.error(request, "No student found with that index number.")
                return render(request, 'exams_office_upload.html', {'form': form})

            # Create a new TranscriptRequest
            t_type = form.cleaned_data.get('transcript_type')
            tr = TranscriptRequest.objects.create(
                student=student,
                transcript_type=t_type,
                reference_code=f"REF-{uuid.uuid4().hex[:8].upper()}",
                payment_made=True,
                amount=TranscriptType.objects.filter(type=t_type).first().price
                if TranscriptType.objects.filter(type=t_type).exists()
                else 0,
            )

            uploaded_pdf = form.cleaned_data.get('transcript_file')
            transcript_obj, _ = Transcript.objects.get_or_create(transcript_request=tr)
            faculty_registrar = form.cleaned_data.get('faculty_registrar')

            # ✅ Step 1: Generate cover page (like your screenshot)
            base_url = request.build_absolute_uri('/')
            if t_type == 'unofficial':
                _, cover_pdf = generate_unofficial_transcript_pdf(tr, faculty_registrar, base_url=base_url)
            else:
                _, cover_pdf = generate_official_transcript_pdf(
                    tr, include_registrar=True, include_vc=True, base_url=base_url
                )

            # ✅ Step 2: Merge cover page + uploaded transcript (if any)
            merged_writer = PdfWriter()

            # Add the cover page first
            cover_reader = PdfReader(BytesIO(cover_pdf))
            for page in cover_reader.pages:
                merged_writer.add_page(page)

            # Add the uploaded transcript pages
            if uploaded_pdf:
                if hasattr(uploaded_pdf, 'temporary_file_path'):
                    with open(uploaded_pdf.temporary_file_path(), 'rb') as uf:
                        upload_reader = PdfReader(uf)
                        for page in upload_reader.pages:
                            merged_writer.add_page(page)
                else:
                    uploaded_pdf.seek(0)
                    upload_reader = PdfReader(uploaded_pdf)
                    for page in upload_reader.pages:
                        merged_writer.add_page(page)

            # ✅ Step 3: Save final merged PDF to transcript_obj.file
            merged_output = BytesIO()
            merged_writer.write(merged_output)
            merged_output.seek(0)

            filename = f"{tr.reference_code}_{t_type}_final.pdf"
            transcript_obj.file.save(filename, ContentFile(merged_output.read()), save=True)

            # ✅ Step 4: Record payment, verification, approval
            payment, _ = Payment.objects.get_or_create(transcript_request=tr)
            payment.officer_name = request.user.get_full_name() or request.user.username
            payment.cleared = True
            payment.notes = (payment.notes or '') + "\nManual upload by exams office."
            payment.save()

            ver_obj, _ = TranscriptVerification.objects.update_or_create(
                transcript=transcript_obj,
                defaults={
                    'barcode': f"VER-{transcript_obj.transcript_id.hex[:12].upper()}",
                    'verified': True,
                    'date_verified': timezone.now(),
                },
            )

            TranscriptApproval.objects.update_or_create(
                transcript=transcript_obj,
                defaults={
                    'approved': False,
                    'remarks': form.cleaned_data.get('remarks', ''),
                    'approved_by': '',
                    'date_approved': None,
                },
            )

            # ✅ Step 5: Create and forward status
            TranscriptStatus.objects.create(
                transcript_request=tr,
                stage='exams_office',
                updated_by=request.user.get_full_name() or request.user.username,
                remarks=f"Manual upload by exams office. {form.cleaned_data.get('remarks', '')}",
            )

            TranscriptStatus.objects.create(
                transcript_request=tr,
                stage='registrar',
                updated_by=request.user.get_full_name() or request.user.username,
                remarks=f"Forwarded from exams office for final approval. {form.cleaned_data.get('remarks', '')}",
            )

            messages.success(
                request,
                f"Transcript uploaded with cover page and forwarded to registrar. Reference: {tr.reference_code}",
            )
            return redirect('staff_dashboard')

    else:
        from .forms import ExamsOfficeUploadForm
        form = ExamsOfficeUploadForm()

    return render(request, 'exams_office_upload.html', {'form': form, 'manual': True})


@login_required
def exams_office_approve_disapprove(request, pk):
    """Exams office approve/disapprove mechanism"""
    transcript_request = get_object_or_404(TranscriptRequest, pk=pk)
    
    staff = StaffProfile.objects.filter(user=request.user, role='exams_office').first()
    if not staff:
        messages.error(request, "Only Exams Office staff can approve/disapprove transcripts.")
        return redirect('staff_dashboard')
    
    transcript, created = Transcript.objects.get_or_create(
        transcript_request=transcript_request,
        defaults={'generated_by': staff.user.get_full_name() or staff.user.username}
    )
    
    current_approval = getattr(transcript, 'approval', None)
    is_currently_approved = current_approval and current_approval.approved if current_approval else False
    
    if request.method == 'POST':
        action = request.POST.get('action')
        remarks = request.POST.get('remarks', '').strip()
        
        if action in ['approve', 'disapprove']:
            if action == 'approve':
                # Check if this is a reapproval of a previously disapproved transcript
                latest_status = transcript_request.statuses.first()
                is_reapproval = latest_status and latest_status.stage == 'disapproved'
                
                # Forward to registrar for final approval
                TranscriptStatus.objects.create(
                    transcript_request=transcript_request,
                    stage='registrar',
                    updated_by=staff.user.get_full_name() or staff.user.username,
                    remarks=f"{'Reapproved' if is_reapproval else 'Approved'} by exams office and forwarded to registrar. {remarks}"
                )
                
                if is_reapproval:
                    messages.success(request, "Transcript reapproved and forwarded to registrar for final approval!")
                else:
                    messages.success(request, "Transcript approved and forwarded to registrar for final approval!")
            else:
                # Disapprove and mark as rejected
                approval, approval_created = TranscriptApproval.objects.get_or_create(
                    transcript=transcript,
                    defaults={
                        'approved': False,
                        'approved_by': staff.user.get_full_name() or staff.user.username,
                        'remarks': remarks,
                        'date_approved': timezone.now()
                    }
                )
                
                if not approval_created:
                    approval.approved = False
                    approval.approved_by = staff.user.get_full_name() or staff.user.username
                    approval.remarks = remarks
                    approval.date_approved = timezone.now()
                    approval.save()
                
                from .models import TranscriptApprovalHistory
                TranscriptApprovalHistory.objects.create(
                    transcript_request=transcript_request,
                    action='disapproved',
                    approved_by=staff.user.get_full_name() or staff.user.username,
                    remarks=remarks
                )
                
                TranscriptStatus.objects.create(
                    transcript_request=transcript_request,
                    stage='disapproved',
                    updated_by=staff.user.get_full_name() or staff.user.username,
                    remarks=remarks or f"Transcript disapproved by Exams Office"
                )
                
                messages.success(request, "Transcript disapproved successfully!")
            
            return redirect('request_detail', pk=pk)
    
    approval_history = transcript_request.approval_history.all()[:10]
    
    # Check if this is a disapproved transcript that can be reapproved
    latest_status = transcript_request.statuses.first()
    is_disapproved = latest_status and latest_status.stage == 'disapproved'
    
    return render(request, 'exams_office_approve_disapprove.html', {
        'transcript_request': transcript_request,
        'transcript': transcript,
        'current_approval': current_approval,
        'is_currently_approved': is_currently_approved,
        'is_disapproved': is_disapproved,
        'approval_history': approval_history,
        'staff': staff
    })




def student_list(request):
    students = StudentProfile.objects.select_related('user').all().order_by('index_number')
    return render(request, "student_list.html", {"students": students})


def student_detail(request, pk):
    student = get_object_or_404(Student, pk=pk)
    requests = student.requests.all()
    return render(request, 'student_detail.html', {
        'student': student,
        'requests': requests
    })


import uuid
import requests
from django.shortcuts import render, redirect
from django.contrib import messages
from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from .models import TranscriptRequest, TranscriptStatus, TranscriptType, StudentProfile
from .forms import TranscriptRequestForm

@login_required
def request_transcript(request):
    profile = StudentProfile.objects.select_related('student').filter(user=request.user).first()
    if not profile or not profile.student:
        messages.error(request, "No student record found for your account.")
        return redirect('student_dashboard')

    student = profile.student
    existing_request = TranscriptRequest.objects.filter(student=student, payment_made=False).first()

    # Fetch transcript prices from DB or use defaults
    try:
        official_price = TranscriptType.objects.get(type='official').price
        unofficial_price = TranscriptType.objects.get(type='unofficial').price
    except TranscriptType.DoesNotExist:
        official_price = 100.00
        unofficial_price = 50.00

    amount = 0
    if existing_request:
        if existing_request.transcript_type == 'official':
            amount = int(official_price * 100)  
        elif existing_request.transcript_type == 'unofficial':
            amount = int(unofficial_price * 100)

    if request.method == "POST" and not request.POST.get("reference"):
        form = TranscriptRequestForm(request.POST, instance=existing_request)
        if 'student' in form.fields:
            form.fields.pop('student')

        if form.is_valid():
            transcript_request = form.save(commit=False)
            transcript_request.student = student
            transcript_request.payment_made = False
            transcript_request.reference_code = f"REF-{uuid.uuid4().hex[:8].upper()}"
            if transcript_request.transcript_type == 'official':
                transcript_request.amount = official_price
            else:
                transcript_request.amount = unofficial_price

            transcript_request.save()

            messages.success(request, "Transcript type selected. Please proceed with payment.")
            return redirect('request_transcript')
        else:
            messages.error(request, "Please correct the errors below.")
    elif request.method == "POST" and request.POST.get("reference"):
        reference = request.POST.get("reference")
        headers = {"Authorization": f"Bearer {settings.PAYSTACK_SECRET_KEY}"}
        verify_url = f"https://api.paystack.co/transaction/verify/{reference}"
        res = requests.get(verify_url, headers=headers).json()

        if res.get("status") and res["data"]["status"] == "success":
            existing_request = TranscriptRequest.objects.filter(student=student, payment_made=False).last()
            if existing_request:
                existing_request.payment_made = True
                existing_request.payment_reference = reference
                existing_request.save()

                TranscriptStatus.objects.create(
                    transcript_request=existing_request,
                    stage='pending',
                    updated_by='System Auto',
                    remarks='Transcript request submitted after payment.'
                )

                messages.success(request, "Payment successful! Your transcript request has been submitted.")
                return redirect('student_detail', pk=student.id)
        else:
            messages.error(request, "Payment verification failed. Please try again.")

    else:
        form = TranscriptRequestForm(instance=existing_request) if existing_request else TranscriptRequestForm()
        if 'student' in form.fields:
            form.fields.pop('student')

    return render(request, "request_transcript.html", {
        "student": student,
        "form": form,
        "existing_request": existing_request,
        "paystack_public_key": settings.PAYSTACK_PUBLIC_KEY,
        "official_price": official_price,
        "unofficial_price": unofficial_price,
        "amount": amount, 
    })


# views.py
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
import requests
from django.conf import settings

@csrf_exempt
@login_required
def init_transcript_payment(request):
    if request.method == "POST":
        amount = 100 * 100 
        email = request.user.email

        headers = {
            "Authorization": f"Bearer {settings.PAYSTACK_SECRET_KEY}",
            "Content-Type": "application/json"
        }
        data = {
            "email": email,
            "amount": amount,
            "callback_url": request.build_absolute_uri('/transcript/payment/verify/')
        }

        response = requests.post("https://api.paystack.co/transaction/initialize",
                                 headers=headers, json=data)
        res_data = response.json()

        if res_data.get('status'):
            return JsonResponse({
                "authorization_url": res_data['data']['authorization_url'],
                "reference": res_data['data']['reference']
            })
        return JsonResponse({"error": "Payment initialization failed"}, status=400)


@login_required
def verify_transcript_payment(request):
    reference = request.GET.get('reference')
    if not reference:
        messages.error(request, "Missing payment reference.")
        return redirect('request_transcript')

    headers = {"Authorization": f"Bearer {settings.PAYSTACK_SECRET_KEY}"}
    response = requests.get(f"https://api.paystack.co/transaction/verify/{reference}", headers=headers)
    res_data = response.json()

    if res_data.get('status') and res_data['data']['status'] == 'success':
        request.session['payment_made'] = True
        messages.success(request, "Payment successful! You can now submit your transcript request.")
        return redirect('request_transcript')
    else:
        messages.error(request, "Payment verification failed. Try again.")
        return redirect('request_transcript')





def safe_str(value):
    if value is None:
        return ""
    if isinstance(value, bytes):
        return value.decode("utf-8", errors="ignore")
    try:
        return unicodedata.normalize("NFKD", str(value)).encode("ascii", "ignore").decode("utf-8", "ignore")
    except Exception:
        return str(value).encode("utf-8", "ignore").decode("utf-8", "ignore")


def safe_send_mail(subject, message, recipient):
    try:
        send_mail(
            safe_str(subject),
            safe_str(message),
            settings.EMAIL_HOST_USER,
            [safe_str(recipient)],
            fail_silently=True,
        )
    except Exception as e:
        print(f"Email send error (ignored): {e}")


def get_student_email(student):
    try:
        if getattr(student, 'user', None) and getattr(student.user, 'email', None):
            return student.user.email
        sp = StudentProfile.objects.select_related('user').filter(student=student).first()
        if sp and getattr(sp.user, 'email', None):
            return sp.user.email
    except Exception:
        pass
    return None


def email_student(subject, message_body, student, attachment=None):
    recipient = get_student_email(student)
    if not recipient:
        return False
    try:
        msg = EmailMultiAlternatives(
            subject=safe_str(subject),
            body=safe_str(message_body),
            from_email=settings.EMAIL_HOST_USER,
            to=[safe_str(recipient)],
        )
        if attachment:
            file_name, content_bytes, mime_type = attachment
            msg.attach(file_name, content_bytes, mime_type or 'application/octet-stream')
        msg.send(fail_silently=True)
        return True
    except Exception as e:
        print(f"Email to student failed: {e}")
        return False



@login_required
def update_payment(request, pk):
    transcript_request = get_object_or_404(TranscriptRequest, pk=pk)
    staff = StaffProfile.objects.filter(user=request.user, role='accounts_office').first()
    if not staff:
        messages.error(request, "Only Accounts Office staff can verify or process payments.")
        return redirect('staff_dashboard')
    if transcript_request.transcript_type == 'unofficial':
        if request.method == 'POST':
            action = request.POST.get('action')
            remarks = request.POST.get('remarks', '')
            if action == 'forward_exams':
                transcript_request.payment_made = True
                transcript_request.save(update_fields=['payment_made'])

                TranscriptStatus.objects.create(
                    transcript_request=transcript_request,
                    stage='exams_office',
                    updated_by=request.user.get_full_name() or request.user.username,
                    remarks=f"Forwarded to Exams Office for unofficial transcript generation. {remarks}"
                )
                student = transcript_request.student
                try:
                    try:
                        from .views import email_student
                    except Exception:
                        email_student = globals().get('email_student')

                    message_body = (
                        f"Dear {student.name},\n\nYour unofficial transcript request (Reference: {transcript_request.reference_code}) has been forwarded to the Exams Office for generation.\n\nBest regards,\nAcademic Records Office"
                    )

                    email_sent = False
                    try:
                        if callable(email_student):
                            email_sent = bool(email_student("Transcript Forwarded to Exams Office", message_body, student))
                    except Exception:
                        email_sent = False

                    if not email_sent:
                        recipient = None
                        try:
                            recipient = getattr(getattr(student, 'user', None), 'email', None)
                        except Exception:
                            recipient = None

                        if not recipient:
                            try:
                                sp = StudentProfile.objects.select_related('user').filter(student=student).first()
                                if sp and getattr(sp.user, 'email', None):
                                    recipient = sp.user.email
                            except Exception:
                                recipient = None

                        if recipient:
                            try:
                                send_mail(
                                    "Transcript Forwarded to Exams Office",
                                    message_body,
                                    settings.DEFAULT_FROM_EMAIL,
                                    [recipient],
                                    fail_silently=True,
                                )
                            except Exception as e:
                                print(f"Email (forward to exams fallback) failed: {e}")
                except Exception as e:
                    print(f"Email (forward to exams) failed: {e}")

                messages.success(request, "Forwarded to Exams Office for transcript generation.")
                return redirect('request_detail', pk=transcript_request.id)

        return render(request, 'clear_student.html', {
            'transcript_request': transcript_request,
            'unofficial_mode': True,
        })


    if request.method == 'POST':
        form = FeeClearanceForm(request.POST, request.FILES)
        if form.is_valid():
            clearance, _ = FeeClearance.objects.get_or_create(transcript_request=transcript_request)
            clearance.cleared = form.cleaned_data.get('cleared')
            clearance.owes = form.cleaned_data.get('owes')
            clearance.amount_owed = form.cleaned_data.get('amount_owed')
            clearance.remarks = form.cleaned_data.get('remarks')
            clearance.officer_name = request.user.get_full_name() or request.user.username
            invoice_file = form.cleaned_data.get('invoice_file') or request.FILES.get('invoice_file')
            if invoice_file:
                clearance.invoice_file = invoice_file
            clearance.save()

            student = transcript_request.student
            recipient = getattr(student.user, 'email', None)

            if clearance.cleared:
                transcript_request.payment_made = True
                transcript_request.save(update_fields=['payment_made'])

                TranscriptStatus.objects.create(
                    transcript_request=transcript_request,
                    stage='accounts_office',
                    updated_by=clearance.officer_name,
                    remarks='Payment verified and forwarded to Registrar for official transcript generation.'
                )

                try:
                    message_body = (
                        f"Dear {student.name},\n\nYour payment has been verified by the Accounts Office. "
                        f"Your request has been forwarded to the Registrar for official transcript generation.\n\n"
                        f"Reference: {transcript_request.reference_code}\n\nBest regards,\nAcademic Records Office"
                    )
                    email_sent = False
                    try:
                        email_sent = bool(email_student("Transcript Payment Verified", message_body, student))
                    except Exception:
                        email_sent = False
                    if not email_sent:
                        recipient = None
                        try:
                            recipient = getattr(getattr(student, 'user', None), 'email', None)
                        except Exception:
                            recipient = None

                        if not recipient:
                            try:
                                sp = StudentProfile.objects.select_related('user').filter(student=student).first()
                                if sp and getattr(sp.user, 'email', None):
                                    recipient = sp.user.email
                            except Exception:
                                recipient = None

                        if recipient:
                            try:
                                send_mail(
                                    "Transcript Payment Verified",
                                    message_body,
                                    settings.DEFAULT_FROM_EMAIL,
                                    [recipient],
                                    fail_silently=True,
                                )
                            except Exception as e:
                                print(f"Email (payment verified fallback) failed: {e}")

                except Exception as e:
                    print(f"Email (payment verified) failed: {e}")

                messages.success(request, "Payment verified and forwarded to Registrar.")

            else:
                transcript_request.payment_made = False
                transcript_request.save(update_fields=['payment_made'])

                TranscriptStatus.objects.create(
                    transcript_request=transcript_request,
                    stage='accounts_office',
                    updated_by=clearance.officer_name,
                    remarks='Payment disapproved — student owes outstanding fees.'
                )

                if recipient:
                    attachment = None
                    try:
                        if clearance.invoice_file and hasattr(clearance.invoice_file, 'path'):
                            import mimetypes, os
                            file_path = clearance.invoice_file.path
                            file_name = os.path.basename(file_path)
                            mime_type, _ = mimetypes.guess_type(file_path)
                            with open(file_path, 'rb') as fh:
                                file_bytes = fh.read()
                            attachment = (file_name, file_bytes, mime_type or 'application/pdf')
                    except Exception:
                        attachment = None

                    message_body = (
                        f"Dear {student.name},\n\nYour transcript payment has been disapproved because you still owe fees. "
                        f"Please find the attached invoice or contact the Accounts Office to resolve this issue.\n\n"
                        f"Reference: {transcript_request.reference_code}\n\nBest regards,\nAcademic Records Office"
                    )
                    try:
                        from .views import email_student
                    except Exception:
                        email_student = globals().get('email_student')
                    if attachment:
                        email_student("Transcript Payment Rejected - Outstanding Fees", message_body, student, attachment=attachment)
                    else:
                        email_student("Transcript Payment Rejected - Outstanding Fees", message_body, student)

                messages.warning(request, "Student owes fees. Transcript disapproved and student notified.")

            return redirect('request_detail', pk=transcript_request.id)

    else:
        form = FeeClearanceForm()

    return render(request, 'clear_student.html', {
        'form': form,
        'transcript_request': transcript_request,
    })




def update_status(request, pk):
    transcript_request = get_object_or_404(TranscriptRequest, pk=pk)
    staff = StaffProfile.objects.filter(user=request.user, role='registrar').first()
    if not staff:
        messages.error(request, "Only Registrar can update status.")
        return redirect('staff_dashboard')
    if request.method == 'POST':
        form = TranscriptStatusForm(request.POST)
        if form.is_valid():
            status = form.save(commit=False)
            status.transcript_request = transcript_request
            status.updated_on = timezone.now()
            status.updated_by = staff.user.get_full_name() or staff.user.username
            status.save()
            messages.success(request, "Status updated successfully!")
            return redirect('request_detail', pk=transcript_request.id)
    else:
        form = TranscriptStatusForm()
    return render(request, 'update_status.html', {
        'form': form,
        'transcript_request': transcript_request
    })






from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.core.mail import send_mail
from django.conf import settings
from django.contrib.auth.decorators import login_required
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from django.core.files.base import ContentFile
from .models import TranscriptRequest, TranscriptStatus, Transcript


def render_transcript_pdf_bytes(transcript_request, request):
    from io import BytesIO
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from django.contrib.staticfiles import finders
    import os

    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    student = transcript_request.student
    registrar_obj = None
    try:
        import re
        st = transcript_request.statuses.filter(stage='exams_office').first()
        if st and st.remarks:
            m = re.search(r'Registrar used:\s*(.+)', st.remarks)
            if m:
                reg_name = m.group(1).strip()
                from .models import FacultyRegistrar
                registrar_obj = FacultyRegistrar.objects.filter(name__iexact=reg_name).first()
    except Exception:
        registrar_obj = None

    if transcript_request.transcript_type == 'unofficial':
        p.setFont("Helvetica-Bold", 18)
        p.drawCentredString(width / 2, height - 80, "Unofficial Transcript")
    else:
        p.setFont("Helvetica-Bold", 18)
        p.drawCentredString(width / 2, height - 80, "OFFICIAL TRANSCRIPT")

    p.setFont("Helvetica", 12)
    y = height - 140
    details = [
        f"Name: {student.name}",
        f"Index Number: {student.index_number}",
        f"Program: {student.program.name}",
        f"Department: {student.department.department}",
        f"Transcript Type: {transcript_request.get_transcript_type_display()}",
        f"Reference Code: {transcript_request.reference_code}",
        f"Date Requested: {transcript_request.date_requested.strftime('%Y-%m-%d %H:%M')}",
    ]
    for line in details:
        p.drawString(100, y, line)
        y -= 20

    p.line(80, y - 10, width - 80, y - 10)

    sig_y = 120
    sig_w = 180
    sig_h = 60
    center_x = width / 2
    sig_x = center_x - (sig_w / 2)

    if transcript_request.transcript_type == 'unofficial':
        try:
            if registrar_obj and getattr(registrar_obj, 'signature', None):
                storage = getattr(registrar_obj.signature, 'storage', None)
                sig_name = getattr(registrar_obj.signature, 'name', None)
                sig_bytes = None
                if storage and sig_name:
                    try:
                        with storage.open(sig_name, 'rb') as fh:
                            sig_bytes = fh.read()
                    except Exception:
                        sig_bytes = None

                if sig_bytes:
                    from reportlab.lib.utils import ImageReader
                    img = ImageReader(BytesIO(sig_bytes))
                    p.drawImage(img, sig_x, sig_y, width=sig_w, height=sig_h, preserveAspectRatio=True, mask='auto')
                    p.setFont("Helvetica-Bold", 10)
                    p.drawCentredString(center_x, sig_y - 14, registrar_obj.name)
                    p.setFont("Helvetica", 9)
                    faculty_label = getattr(registrar_obj, 'faculty_name', '')
                    if faculty_label:
                        p.drawCentredString(center_x, sig_y - 28, f"{faculty_label} (Registrar)")
                    else:
                        p.drawCentredString(center_x, sig_y - 28, "(Registrar)")
                else:
                    p.setFont("Helvetica", 10)
                    if registrar_obj:
                        p.drawCentredString(center_x, sig_y + sig_h / 2 - 6,
                                            f"{registrar_obj.name} - {getattr(registrar_obj, 'faculty_name', '')} (Registrar)")
                    else:
                        p.drawCentredString(center_x, sig_y + sig_h / 2 - 6, "Registrar (not selected)")
        except Exception:
            pass
    else:
        # Try to use uploaded StaffProfile.signature for Registrar and VC; fall back to static images
        from .models import StaffProfile
        x_positions = [180, 360]
        try:
            reg_profile = StaffProfile.objects.filter(role='registrar').select_related('user').first()
            drawn = False
            if reg_profile and getattr(reg_profile, 'signature', None):
                try:
                    storage = getattr(reg_profile.signature, 'storage', None)
                    sig_name = getattr(reg_profile.signature, 'name', None)
                    if storage and sig_name:
                        with storage.open(sig_name, 'rb') as fh:
                            sig_bytes = fh.read()
                            from reportlab.lib.utils import ImageReader
                            img = ImageReader(BytesIO(sig_bytes))
                            p.drawImage(img, x_positions[0], sig_y, width=120, height=60, preserveAspectRatio=True, mask='auto')
                            drawn = True
                except Exception:
                    drawn = False

            if not drawn:
                reg_img = finders.find('images/registrar.jpg') or os.path.join(os.path.dirname(__file__), 'images', 'registrar.jpg')
                if reg_img and os.path.exists(reg_img):
                    try:
                        img = ImageReader(reg_img)
                        p.drawImage(img, x_positions[0], sig_y, width=120, height=60, preserveAspectRatio=True, mask='auto')
                    except Exception:
                        pass

            p.setFont("Helvetica-Bold", 10)
            reg_name = (reg_profile.user.get_full_name() if getattr(reg_profile, 'user', None) else None) or 'Registrar'
            p.drawCentredString(x_positions[0] + 60, sig_y - 16, reg_name)
            p.setFont("Helvetica", 9)
            p.drawCentredString(x_positions[0] + 60, sig_y - 30, 'Registrar')
        except Exception:
            pass

        try:
            vc_profile = StaffProfile.objects.filter(role='vice_chancellor').select_related('user').first()
            drawn_vc = False
            if vc_profile and getattr(vc_profile, 'signature', None):
                try:
                    storage = getattr(vc_profile.signature, 'storage', None)
                    sig_name = getattr(vc_profile.signature, 'name', None)
                    if storage and sig_name:
                        with storage.open(sig_name, 'rb') as fh:
                            sig_bytes = fh.read()
                            from reportlab.lib.utils import ImageReader
                            img = ImageReader(BytesIO(sig_bytes))
                            p.drawImage(img, x_positions[1], sig_y, width=120, height=60, preserveAspectRatio=True, mask='auto')
                            drawn_vc = True
                except Exception:
                    drawn_vc = False

            if not drawn_vc:
                vc_img = finders.find('images/vice_chancellor.png') or os.path.join(os.path.dirname(__file__), 'images', 'vice_chancellor.png')
                if vc_img and os.path.exists(vc_img):
                    try:
                        img = ImageReader(vc_img)
                        p.drawImage(img, x_positions[1], sig_y, width=120, height=60, preserveAspectRatio=True, mask='auto')
                    except Exception:
                        pass

            p.setFont("Helvetica-Bold", 10)
            vc_name = (vc_profile.user.get_full_name() if getattr(vc_profile, 'user', None) else None) or 'Vice Chancellor'
            p.drawCentredString(x_positions[1] + 60, sig_y - 16, vc_name)
            p.setFont("Helvetica", 9)
            p.drawCentredString(x_positions[1] + 60, sig_y - 30, 'Vice Chancellor')
        except Exception:
            pass

    try:
        from reportlab.graphics.shapes import Drawing
        from reportlab.graphics.barcode.qr import QrCodeWidget
        from reportlab.graphics import renderPDF
        import urllib.parse, re

        ref = transcript_request.reference_code or ''
        ref_clean = re.sub(r'(?i)^ref[-_]?','', ref)
        index = getattr(student, 'index_number', '')
        code = f"{index}{ref_clean}"

        base_url = request.build_absolute_uri('/')[:-1]
        encoded = urllib.parse.quote(code, safe='')
        verify_url = f"{base_url}/verify/{encoded}/"
        qr_size = 120
        qr_x = (width - qr_size) / 2
        qr_y = (height - qr_size) / 2
        qr = QrCodeWidget(verify_url)
        d = Drawing(qr_size, qr_size)
        d.add(qr)
        renderPDF.draw(d, p, qr_x, qr_y)

        text_y = qr_y - 12
        p.setFont("Helvetica", 10)
        p.drawCentredString(width / 2, text_y, "Scan the QR code above to verify this transcript.")
        text_y -= 14
        p.setFont("Helvetica", 9)
        p.drawCentredString(width / 2, text_y, f"Or visit: {base_url}/verify/")
        text_y -= 14
        p.setFont("Helvetica-Oblique", 8)
        p.drawCentredString(width / 2, text_y, f"You can also verify manually using the Reference Code at: {base_url}/verify/")

    except Exception:
        pass

    p.showPage()
    p.save()
    buffer.seek(0)
    return buffer.getvalue()





@login_required
def registrar_action(request, pk):
    transcript_request = get_object_or_404(TranscriptRequest, id=pk)
    staff = request.user

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'approve':
            TranscriptStatus.objects.create(
                transcript_request=transcript_request,
                stage='approved',
                remarks=f"Approved by {staff.get_full_name() or staff.username}",
                updated_by=staff.get_full_name() or staff.username,
            )

            transcript_request.status = 'approved'
            transcript_request.save()
            student = transcript_request.student
            try:
                transcript_obj, created = Transcript.objects.get_or_create(
                    transcript_request=transcript_request,
                    defaults={'generated_by': staff.get_full_name() or staff.username},
                )

                if not transcript_obj.file:
                    try:
                        pdf_bytes = render_transcript_pdf_bytes(transcript_request, request)
                        safe_name = student.name.replace(" ", "_").replace("/", "_")
                        file_name = f"{safe_name}_Transcript.pdf"
                        transcript_obj.file.save(file_name, ContentFile(pdf_bytes))
                    except Exception as e:
                        print(f"Failed to build/save final PDF on approval: {e}")

                transcript_obj.registrar_signature = True
                transcript_obj.save()
            except Exception as e:
                print(f"[Registrar Approval Error] {e}")
            try:
                subject = "Transcript Approved"
                message = (
                    f"Dear {student.name},\n\n"
                    f"Your transcript request ({transcript_request.reference_code}) has been approved by the Registrar. "
                    f"You can download your transcript from your dashboard.\n\n"
                    f"Best regards,\nAcademic Records Office"
                )
                email_sent = False
                try:
                    attachment = None
                    try:
                        if transcript_obj and getattr(transcript_obj, 'file', None):
                            tf = transcript_obj.file
                            tf.open('rb')
                            file_bytes = tf.read()
                            tf.close()
                            import os, mimetypes
                            filename = os.path.basename(tf.name) or f"{student.name.replace(' ', '_')}_Transcript.pdf"
                            mime_type, _ = mimetypes.guess_type(filename)
                            attachment = (filename, file_bytes, mime_type or 'application/pdf')
                    except Exception:
                        attachment = None

                    email_sent = email_student(subject, message, student, attachment=attachment)
                except Exception:
                    email_sent = False

                if not email_sent:
                    recipient = getattr(student, 'email', None) or getattr(getattr(student, 'user', None), 'email', None)
                    if recipient:
                        try:
                            if attachment:
                                try:
                                    msg = EmailMultiAlternatives(subject, message, settings.DEFAULT_FROM_EMAIL, [recipient])
                                    msg.attach(attachment[0], attachment[1], attachment[2])
                                    msg.send(fail_silently=True)
                                except Exception:
                                    send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [recipient], fail_silently=True)
                            else:
                                send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [recipient], fail_silently=True)
                        except Exception:
                            pass
            except Exception as e:
                print(f"[Email Notification Error] {e}")

            messages.success(request, "Transcript approved successfully.")
            return redirect("staff_dashboard")

        elif action == 'reject':
            comments = request.POST.get('comments', '').strip()
            attach = request.FILES.get('attachment') if request.FILES else None
            TranscriptStatus.objects.create(
                transcript_request=transcript_request,
                stage='rejected',
                remarks=comments or f"Disapproved by {staff.get_full_name() or staff.username}",
                updated_by=staff.get_full_name() or staff.username,
                attachment=attach
            )

            transcript_request.status = 'rejected'
            transcript_request.save()

            try:
                student = transcript_request.student
                subject = "Transcript Request Disapproved"
                message = (
                    f"Dear {student.name},\n\n"
                    f"Your transcript request ({transcript_request.reference_code}) has been disapproved by the Registrar.\n\n"
                    f"Remarks: {comments or 'No remarks provided.'}\n\n"
                    f"Best regards,\nAcademic Records Office"
                )
                email_sent = False
                try:
                    attachment = None
                    try:
                        latest = transcript_request.statuses.order_by('-updated_on').first()
                        if latest and getattr(latest, 'attachment', None):
                            a = latest.attachment
                            a.open('rb')
                            a_bytes = a.read()
                            a.close()
                            import os, mimetypes
                            a_name = os.path.basename(a.name) or f"rejection_{transcript_request.reference_code}.pdf"
                            a_type, _ = mimetypes.guess_type(a_name)
                            attachment = (a_name, a_bytes, a_type or 'application/octet-stream')
                    except Exception:
                        attachment = None

                    email_sent = email_student(subject, message, student, attachment=attachment)
                except Exception:
                    email_sent = False

                if not email_sent:
                    recipient = getattr(student, 'email', None) or getattr(getattr(student, 'user', None), 'email', None)
                    if recipient:
                        try:
                            if attachment:
                                try:
                                    msg = EmailMultiAlternatives(subject, message, settings.DEFAULT_FROM_EMAIL, [recipient])
                                    msg.attach(attachment[0], attachment[1], attachment[2])
                                    msg.send(fail_silently=True)
                                except Exception:
                                    send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [recipient], fail_silently=True)
                            else:
                                send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [recipient], fail_silently=True)
                        except Exception:
                            pass
            except Exception as e:
                print(f"[Disapproval Email Error] {e}")

            messages.warning(request, "Transcript disapproved successfully.")
            return redirect("staff_dashboard")

    context = {
        'transcript_request': transcript_request,
    }
    return render(request, 'registrar_action.html', context)


@login_required
def registrar_upload_transcript(request, pk):
    """Registrar uploads a transcript PDF (manual). Handles payment breakdown and generates the stored official/unofficial PDF."""
    # Only registrar staff may perform this
    staff = StaffProfile.objects.filter(user=request.user, role='registrar').first()
    if not staff:
        messages.error(request, "Only Registrar staff can upload transcripts here.")
        return redirect('staff_dashboard')

    transcript_request = get_object_or_404(TranscriptRequest, pk=pk)

    if request.method == 'POST':
        form = RegistrarUploadForm(request.POST, request.FILES)
        if form.is_valid():
            # Save uploaded file to Transcript.uploaded_file and create/attach Transcript object
            uploaded = form.cleaned_data.get('transcript_file')
            momo_ref = None
            t_type = form.cleaned_data.get('transcript_type')
            faculty_registrar = form.cleaned_data.get('faculty_registrar')

            transcript_obj, _ = Transcript.objects.get_or_create(transcript_request=transcript_request)
            if uploaded:
                transcript_obj.uploaded_file = uploaded
                transcript_obj.save()

            # Record payment breakdown (create or update Payment)
            payment, _ = Payment.objects.get_or_create(transcript_request=transcript_request)
            payment.amount_received = form.cleaned_data.get('amount_received')
            payment.amount_accounts_office = form.cleaned_data.get('amount_accounts_office')
            payment.amount_superadmin = form.cleaned_data.get('amount_superadmin')
            payment.amount_registrar = form.cleaned_data.get('amount_registrar')
            payment.officer_name = request.user.get_full_name() or request.user.username
            payment.cleared = True
            payment.notes = (payment.notes or '') + f"\nRegistrar uploaded file. Momo Ref: {momo_ref or ''}"
            payment.save()

            # Mark request payment_made and payment_reference
            transcript_request.payment_made = True
            if momo_ref:
                transcript_request.payment_reference = momo_ref
            transcript_request.save()

            # Generate official/unofficial PDF using utils
            base_url = request.build_absolute_uri('/')
            if t_type == 'unofficial':
                transcript_obj, pdf_bytes = generate_unofficial_transcript_pdf(transcript_request, faculty_registrar, base_url=base_url)
            else:
                transcript_obj, pdf_bytes = generate_official_transcript_pdf(transcript_request, include_registrar=True, include_vc=True, base_url=base_url)

            # Save file if utils returned bytes
            if pdf_bytes and transcript_obj:
                # ensure uploaded_file already saved; transcript_obj.file saved by utils
                # Create or update verification record. For official transcripts uploaded by Registrar,
                # mark verified immediately and create approval/status records.
                ver_obj, created = TranscriptVerification.objects.update_or_create(
                    transcript=transcript_obj,
                    defaults={
                        'barcode': f"VER-{transcript_obj.transcript_id.hex[:12].upper()}",
                        'verified': True if t_type == 'official' else False,
                        'date_verified': timezone.now() if t_type == 'official' else None,
                    }
                )

                if t_type == 'official':
                    # create an approval record and set status to approved
                    TranscriptApproval.objects.update_or_create(
                        transcript=transcript_obj,
                        defaults={
                            'approved': True,
                            'approved_by': request.user.get_full_name() or request.user.username,
                            'date_approved': timezone.now(),
                        }
                    )

                    TranscriptStatus.objects.create(
                        transcript_request=transcript_request,
                        stage='approved',
                        remarks=f"Official transcript generated and verified by Registrar. Momo:{momo_ref or ''}",
                        updated_by=request.user.get_full_name() or request.user.username
                    )
                else:
                    # Unofficial: mark registrar stage and note selection
                    TranscriptStatus.objects.create(
                        transcript_request=transcript_request,
                        stage='registrar',
                        remarks=f"Unofficial transcript uploaded by registrar. Faculty Registrar: {getattr(faculty_registrar, 'name', '')}",
                        updated_by=request.user.get_full_name() or request.user.username
                    )

            messages.success(request, "Transcript uploaded and processed.")
            return redirect('request_detail', pk=transcript_request.id)
        else:
            messages.error(request, "Please correct the errors in the form.")
    else:
        # pre-fill student index
        initial = {'student_index': transcript_request.student.index_number if transcript_request and transcript_request.student else ''}
        form = RegistrarUploadForm(initial=initial)

    return render(request, 'registrar_upload.html', {'form': form, 'transcript_request': transcript_request})





from django.shortcuts import render, redirect
from .forms import FacultyRegistrarForm, FacultyChangeForm

def change_faculty_registrar(request, pk):
    """Allow registrar to change faculty for unofficial transcripts"""
    transcript_request = get_object_or_404(TranscriptRequest, pk=pk)
    
    staff = StaffProfile.objects.filter(user=request.user, role='registrar').first()
    if not staff:
        messages.error(request, "Only Registrar can change faculty selection.")
        return redirect('staff_dashboard')
    
    if transcript_request.transcript_type != 'unofficial':
        messages.error(request, "Faculty change is only allowed for unofficial transcripts.")
        return redirect('request_detail', pk=pk)
    
    if request.method == 'POST':
        form = FacultyChangeForm(request.POST)
        if form.is_valid():
            old_faculty = transcript_request.selected_faculty_registrar
            new_faculty = form.cleaned_data['faculty_registrar']
            reason = form.cleaned_data['reason']
            
            transcript_request.selected_faculty_registrar = new_faculty
            transcript_request.save()
            
            TranscriptStatus.objects.create(
                transcript_request=transcript_request,
                stage='registrar',
                updated_by=staff.user.get_full_name() or staff.user.username,
                remarks=f"Faculty changed from {old_faculty} to {new_faculty}. Reason: {reason}"
            )
            
            try:
                from .utils import generate_unofficial_transcript_pdf
                base_url = request.build_absolute_uri('/')[:-1]
                transcript_obj, pdf_bytes = generate_unofficial_transcript_pdf(
                    transcript_request, new_faculty, base_url
                )
                messages.success(request, f"Faculty changed successfully to {new_faculty.name}.")
            except Exception as e:
                messages.warning(request, f"Faculty changed but transcript regeneration failed: {str(e)}")
            
            return redirect('request_detail', pk=pk)
    else:
        form = FacultyChangeForm(initial={
            'faculty_registrar': transcript_request.selected_faculty_registrar
        })
    
    return render(request, 'change_faculty.html', {
        'form': form,
        'transcript_request': transcript_request,
        'current_faculty': transcript_request.selected_faculty_registrar
    })


def registrar_approve_disapprove(request, pk):
    """Improved approve/disapprove mechanism with history tracking"""
    transcript_request = get_object_or_404(TranscriptRequest, pk=pk)
    
    staff = StaffProfile.objects.filter(user=request.user, role='registrar').first()
    if not staff:
        messages.error(request, "Only Registrar can approve/disapprove transcripts.")
        return redirect('staff_dashboard')
    
    transcript, created = Transcript.objects.get_or_create(
        transcript_request=transcript_request,
        defaults={'generated_by': staff.user.get_full_name() or staff.user.username}
    )
    
    current_approval = getattr(transcript, 'approval', None)
    is_currently_approved = current_approval and current_approval.approved if current_approval else False
    
    if request.method == 'POST':
        action = request.POST.get('action')
        remarks = request.POST.get('remarks', '').strip()
        
        if action in ['approve', 'disapprove']:
            if action == 'approve':
                if is_currently_approved:
                    history_action = 'reapproved'
                else:
                    history_action = 'approved'
            else:
                history_action = 'disapproved'
            
            approval, approval_created = TranscriptApproval.objects.get_or_create(
                transcript=transcript,
                defaults={
                    'approved': action == 'approve',
                    'approved_by': staff.user.get_full_name() or staff.user.username,
                    'remarks': remarks,
                    'date_approved': timezone.now()
                }
            )
            
            if not approval_created:
                approval.approved = action == 'approve'
                approval.approved_by = staff.user.get_full_name() or staff.user.username
                approval.remarks = remarks
                approval.date_approved = timezone.now()
                approval.save()
            
            from .models import TranscriptApprovalHistory
            TranscriptApprovalHistory.objects.create(
                transcript_request=transcript_request,
                action=history_action,
                approved_by=staff.user.get_full_name() or staff.user.username,
                remarks=remarks
            )
            
            stage = 'approved' if action == 'approve' else 'rejected'
            TranscriptStatus.objects.create(
                transcript_request=transcript_request,
                stage=stage,
                updated_by=staff.user.get_full_name() or staff.user.username,
                remarks=remarks or f"Transcript {action}d by Registrar"
            )
            
            if action == 'approve':
                try:
                    from .utils import generate_unofficial_transcript_pdf, generate_official_transcript_pdf
                    base_url = request.build_absolute_uri('/')[:-1]
                    
                    if transcript_request.transcript_type == 'unofficial':
                        faculty_registrar = transcript_request.selected_faculty_registrar
                        transcript_obj, pdf_bytes = generate_unofficial_transcript_pdf(
                            transcript_request, faculty_registrar, base_url
                        )
                    else:
                        transcript_obj, pdf_bytes = generate_official_transcript_pdf(
                            transcript_request, include_registrar=True, include_vc=True, base_url=base_url
                        )
                except Exception as e:
                    messages.warning(request, f"Transcript {action}d but generation failed: {str(e)}")
            
            messages.success(request, f"Transcript {action}d successfully!")
            return redirect('request_detail', pk=pk)
    
    approval_history = transcript_request.approval_history.all()[:10]  
    
    return render(request, 'registrar_approve_disapprove.html', {
        'transcript_request': transcript_request,
        'transcript': transcript,
        'current_approval': current_approval,
        'is_currently_approved': is_currently_approved,
        'approval_history': approval_history,
        'staff': staff
    })

def add_faculty_registrar(request):
    if request.method == 'POST':
        form = FacultyRegistrarForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('faculty_registrar_list')  
    else:
        form = FacultyRegistrarForm()
    return render(request, 'faculty_registrar_form.html', {'form': form})



def approve_transcript(request, pk):
    transcript = get_object_or_404(Transcript, pk=pk)
    if request.method == 'POST':
        form = TranscriptApprovalForm(request.POST)
        if form.is_valid():
            approval = form.save(commit=False)
            approval.transcript = transcript
            approval.date_approved = timezone.now()
            approval.save()

            TranscriptStatus.objects.create(
                transcript_request=transcript.transcript_request,
                stage='approved' if approval.approved else 'rejected',
                updated_by=approval.approved_by,
                remarks=approval.remarks or ''
            )
            messages.success(request, "Transcript approval updated successfully!")

            try:
                tr = transcript.transcript_request
                student = tr.student
                if approval.approved:
                    subject = "Your Official Transcript is Approved"
                    message_body = (
                        f"Dear {student.name},\n\n"
                        f"Your official transcript has been approved by the Registrar. The PDF is attached for your records.\n\n"
                        f"Reference: {tr.reference_code}\n\n"
                        f"Best regards,\n"
                        f"Academic Records Office"
                    )
                    try:
                        buffer = BytesIO()
                        p = canvas.Canvas(buffer, pagesize=A4)
                        width, height = A4
                        p.setFont("Helvetica-Bold", 18)
                        p.drawCentredString(width / 2, height - 80, "OFFICIAL TRANSCRIPT")
                        y = height - 140
                        details = [
                            f"Name: {student.name}",
                            f"Index Number: {student.index_number}",
                            f"Program: {student.program.name}",
                            f"Department: {student.department.department}",
                            f"Transcript Type: {tr.get_transcript_type_display()}",
                            f"Reference Code: {tr.reference_code}",
                            f"Date Requested: {tr.date_requested.strftime('%Y-%m-%d %H:%M')}",
                        ]
                        for line in details:
                            p.drawString(100, y, line)
                            y -= 20
                        p.line(80, y - 10, width - 80, y - 10)
                        p.setFont("Helvetica-Oblique", 9)
                        p.drawCentredString(width / 2, 40, "Approved by Registrar")
                        p.showPage()
                        p.save()
                        buffer.seek(0)
                        safe_name = student.name.replace(" ", "_").replace("/", "_")
                        file_name = f"{safe_name}_Transcript.pdf"
                        try:
                            if not hasattr(tr, 'transcript'):
                                # If this TranscriptRequest has no Transcript record, create one
                                transcript_obj, created = Transcript.objects.get_or_create(
                                    transcript_request=tr,
                                    defaults={
                                        'generated_by': approval.approved_by or 'Registrar',
                                    },
                                )
                            else:
                                transcript_obj = tr.transcript

                            # Persist the PDF bytes to the filefield
                            transcript_obj.file.save(file_name, ContentFile(buffer.getvalue()))
                            transcript_obj.registrar_signature = True
                            transcript_obj.save()
                        except Exception as e:
                            print(f"Failed to save approved transcript file: {e}")

                        # Email the student with attachment (fall back handled inside email_student)
                        email_student(subject, message_body, student, attachment=(file_name, buffer.getvalue(), 'application/pdf'))
                    except Exception as e:
                        print(f"Failed to build final PDF for email: {e}")
                        email_student(subject, message_body, student)
                else:
                    subject = "Transcript Rejected"
                    message_body = (
                        f"Dear {student.name},\n\n"
                        f"Your transcript was not approved by the Registrar.\n"
                        f"Remarks: {approval.remarks or 'No remarks provided.'}\n\n"
                        f"Reference: {tr.reference_code}\n\n"
                        f"Please check your dashboard for details.\n\n"
                        f"Best regards,\n"
                        f"Academic Records Office"
                    )
                    email_student(subject, message_body, student)
            except Exception as e:
                print(f"Email from Registrar step failed: {e}")
            return redirect('request_detail', pk=transcript.transcript_request.id)
    else:
        form = TranscriptApprovalForm()
    return render(request, 'approve_transcript.html', {
        'form': form,
        'transcript': transcript
    })


def request_detail(request, pk):
    transcript_request = get_object_or_404(TranscriptRequest, pk=pk)
    statuses = transcript_request.statuses.all()
    # Compute a template-safe boolean for whether the transcript is approved.
    # Doing this in the view avoids calling QuerySet methods with args in templates.
    transcript_approved = False
    try:
        transcript_approved = bool(
            getattr(transcript_request, 'transcript', None) and
            transcript_request.statuses.filter(stage='approved').exists()
        )
    except Exception:
        transcript_approved = False

    # Determine whether the current user is a linked student and whether they may download the invoice
    profile = None
    is_student_user = False
    can_download_invoice = False
    try:
        if request.user.is_authenticated:
            profile = StudentProfile.objects.select_related('student').filter(user=request.user).first()
            is_student_user = bool(profile and profile.student)
            clearance = getattr(transcript_request, 'fee_clearance', None)
            can_download_invoice = bool(profile and profile.student == transcript_request.student and clearance and clearance.invoice_file)
            # Also determine staff role for staff users so templates can show staff-only controls
            staff_role = None
            if request.user.is_staff:
                sp = StaffProfile.objects.filter(user=request.user).first()
                staff_role = sp.role.lower() if sp and getattr(sp, 'role', None) else None
    except Exception:
        # If anything goes wrong determining profile/invoice access, default to False
        is_student_user = False
        can_download_invoice = False

    # Only show the update status control to registrar staff users
    show_update = False
    try:
        show_update = True if locals().get('staff_role', None) == 'registrar' else False
    except Exception:
        show_update = False

    return render(request, 'request_detail_fixed.html', {
        'transcript_request': transcript_request,
        'statuses': statuses,
        'transcript_approved': transcript_approved,
        'is_student_user': is_student_user,
        'can_download_invoice': can_download_invoice,
        'show_update_status': show_update,
        'staff_role': locals().get('staff_role', None),
    })


def add_department(request):
    if request.method == "POST":
        dept_name = request.POST.get("department")
        hod_name = request.POST.get("HoD")
        Department.objects.create(department=dept_name, HoD=hod_name)
        return redirect("department_list")  
    return render(request, "add_department.html")




@login_required
def student_generate_transcript(request, pk):
    transcript_request = get_object_or_404(TranscriptRequest, pk=pk)
    profile = StudentProfile.objects.select_related('student').filter(user=request.user).first()
    if not profile or profile.student != transcript_request.student:
        messages.error(request, "You don't have permission to generate this transcript.")
        return redirect('student_approved_transcripts')

    approved = transcript_request.statuses.filter(stage='approved').exists() or (
        getattr(getattr(transcript_request, 'transcript', None), 'transcript_request', None) and getattr(transcript_request, 'transcript', None)
    )
    if not approved:
        messages.error(request, "This request is not approved yet.")
        return redirect('student_approved_transcripts')

    transcript_obj = getattr(transcript_request, 'transcript', None)
    if transcript_obj and transcript_obj.file:
        messages.info(request, "Transcript PDF is already available.")
        return redirect('student_approved_transcripts')

    try:
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4
        p.setFont("Helvetica-Bold", 18)
        p.drawCentredString(width / 2, height - 80, "OFFICIAL TRANSCRIPT")
        y = height - 140
        student = transcript_request.student
        details = [
            f"Name: {student.name}",
            f"Index Number: {student.index_number}",
            f"Program: {student.program.name}",
            f"Department: {student.department.department}",
            f"Transcript Type: {transcript_request.get_transcript_type_display()}",
            f"Reference Code: {transcript_request.reference_code}",
            f"Date Requested: {transcript_request.date_requested.strftime('%Y-%m-%d %H:%M')}",
        ]
        for line in details:
            p.drawString(100, y, line)
            y -= 20
        p.line(80, y - 10, width - 80, y - 10)
        p.setFont("Helvetica-Oblique", 9)
        p.drawCentredString(width / 2, 40, "Approved by Registrar")
        p.showPage()
        p.save()
        buffer.seek(0)

        safe_name = student.name.replace(" ", "_").replace("/", "_")
        file_name = f"{safe_name}_Transcript.pdf"

        if not transcript_obj:
            transcript_obj, created = Transcript.objects.get_or_create(
                transcript_request=transcript_request,
                defaults={'generated_by': request.user.get_full_name() or request.user.username}
            )

        transcript_obj.file.save(file_name, ContentFile(buffer.getvalue()))
        transcript_obj.registrar_signature = True
        transcript_obj.save()

        messages.success(request, "Transcript PDF generated and saved. It will appear in the list now.")
    except Exception as e:
        print(f"Error generating transcript for student: {e}")
        messages.error(request, "Failed to generate transcript PDF. Contact support.")

    return redirect('student_approved_transcripts')



def department_list(request):
    departments = Department.objects.all().order_by('department') 
    return render(request, "department_list.html", {"departments": departments})

def add_department(request):
    if request.method == "POST":
        dept_name = request.POST.get("department")
        hod_name = request.POST.get("HoD")
        if dept_name and hod_name:
            Department.objects.create(department=dept_name, HoD=hod_name)
        return redirect("department_list")
    return render(request, "add_department.html")


def update_department(request, pk):
    dept = get_object_or_404(Department, pk=pk)
    if request.method == "POST":
        dept.department = request.POST.get("department")
        dept.HoD = request.POST.get("HoD")
        dept.save()
        return redirect("department_list")
    return render(request, "update_department.html", {"dept": dept})

def delete_department(request, pk):
    dept = get_object_or_404(Department, pk=pk)
    if request.method == "POST":
        dept.delete()
        return redirect("department_list")
    return render(request, "delete_department.html", {"dept": dept})




def signup_view(request):
    if request.method == "POST":
        form = CustomSignupForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.is_staff = False
            user.save()
            StudentProfile.objects.create(user=user, index_number=form.cleaned_data["index_number"], phone_number=form.cleaned_data["phone"])
            login(request, user)
            messages.success(request, "Account created successfully!")
            return redirect("student_dashboard")
        else:
            messages.error(request, "Please fix the errors below.")
    else:
        form = CustomSignupForm()
    return render(request, "signup.html", {"form": form})

class DebugPasswordResetConfirmView(auth_views.PasswordResetConfirmView):
    """A subclassed PasswordResetConfirmView that logs UID/token checks for debugging.

    This view should be used in development only to understand why tokens may be rejected.
    """
    def dispatch(self, request, *args, **kwargs):
        uidb64 = kwargs.get('uidb64')
        token = kwargs.get('token')
        try:
            uid = urlsafe_base64_decode(uidb64).decode()
            user = User.objects.filter(pk=uid).first()
            user_email = getattr(user, 'email', None) if user else None
        except Exception:
            uid = None
            user = None
            user_email = None

        try:
            print(f"[DebugPasswordReset] dispatch uidb64={uidb64} uid={uid} email={user_email} token={token}")
            token_ok = bool(user and default_token_generator.check_token(user, token))
            print(f"[DebugPasswordReset] token valid: {token_ok}")
        except Exception as e:
            print(f"[DebugPasswordReset] token check failed: {e}")

        return super().dispatch(request, *args, **kwargs)


def password_reset_request(request):
    """Handles password reset email or student ID submission."""
    if request.method == 'POST':
        if 'email' in request.POST and 'csrfmiddlewaretoken' in request.POST and len(request.POST) <= 3:
            email = request.POST.get('email', '').strip().lower()
            form = PasswordResetRequestForm({'email': email})
        else:
            form = PasswordResetRequestForm(request.POST)

        if form.is_valid():
            raw = form.cleaned_data['email'].strip()
            user = None
            resolved_email = None

            # Determine if it's an email or student index number
            if '@' in raw:
                resolved_email = raw.lower()
                user = User.objects.filter(email__iexact=resolved_email).first()
            else:
                sp = StudentProfile.objects.select_related('user').filter(index_number__iexact=raw).first()
                if sp and getattr(sp, 'user', None):
                    user = sp.user
                    resolved_email = getattr(user, 'email', None)

            if user:
                # Generate and store a reset code
                code = secrets.token_urlsafe(16)
                expires = timezone.now() + timedelta(minutes=15)
                pr = PasswordResetCode.objects.create(user=user, code=code, expires_at=expires)

                print(f"[PasswordReset] Created code id={pr.id}, user={user.email}, code={pr.code}, expires={pr.expires_at}")

                # Identify student if possible
                sp2 = StudentProfile.objects.filter(user=user).first()
                identifier = sp2.index_number if sp2 and getattr(sp2, 'index_number', None) else resolved_email

                # Build reset URL
                confirm_url = request.build_absolute_uri(
                    f"{reverse('app_password_reset_confirm')}?code={urllib.parse.quote_plus(pr.code)}&id={urllib.parse.quote_plus(identifier)}"
                )

                subject = 'Password reset code'
                message = (
                    f"Use this code to reset your password (valid for 15 minutes):\n\n{pr.code}\n\n"
                    f"Or click this link:\n{confirm_url}\n\n"
                    f"If you didn’t request this, ignore this message."
                )

                # Try to send email
                try:
                    send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [resolved_email], fail_silently=True)
                    print(f"[PasswordReset] Email sent to {resolved_email}")
                except Exception as e:
                    print(f"[PasswordReset] Email sending failed: {e}")

                return render(request, 'registration/password_reset_sent.html', {
                    'email': form.cleaned_data.get('email'),
                    'identifier': identifier,
                })

    else:
        form = PasswordResetRequestForm()

    return render(request, 'registration/password_reset_request.html', {'form': form})


def password_reset_confirm(request):
    """Verifies the reset code and allows setting a new password."""
    if request.method == 'POST':
        form = PasswordResetConfirmForm(request.POST)
        if form.is_valid():
            raw_code = form.cleaned_data['code'].strip()
            identifier = (form.cleaned_data.get('email_or_index') or '').strip()

            # Handle URL encoding and decoding variations
            candidates = {raw_code}
            try:
                candidates.add(urllib.parse.unquote(raw_code))
                candidates.add(urllib.parse.unquote_plus(raw_code))
            except Exception:
                pass
            if ' ' in raw_code:
                candidates.add(raw_code.replace(' ', '+'))

            resolved_user = None
            if identifier:
                if '@' in identifier:
                    resolved_user = User.objects.filter(email__iexact=identifier).first()
                else:
                    sp = StudentProfile.objects.filter(index_number__iexact=identifier).select_related('user').first()
                    if sp and getattr(sp, 'user', None):
                        resolved_user = sp.user

            print(f"[PasswordReset] Confirm attempt: identifier='{identifier}', candidates={list(candidates)}")
            if resolved_user:
                print(f"[PasswordReset] Resolved user: {resolved_user.email}")

            # Try to find a matching reset code
            pr = None
            for c in candidates:
                if not c:
                    continue
                qs = PasswordResetCode.objects.filter(code=c, used=False)
                if resolved_user:
                    qs = qs.filter(user=resolved_user)
                pr_candidate = qs.select_related('user').first()
                if pr_candidate:
                    pr = pr_candidate
                    print(f"[PasswordReset] Found code id={pr.id}, user={pr.user.email}, used={pr.used}, expires={pr.expires_at}")
                    break

            # Check code validity
            if pr and pr.is_valid():
                print(f"[PasswordReset] Code valid. Updating password for {pr.user.email}")
                user = pr.user
                new_password = form.cleaned_data['new_password1']
                user.set_password(new_password)
                user.save()
                pr.used = True
                pr.save()
                return render(request, 'registration/password_reset_complete.html')

            # Handle invalid code
            print(f"[PasswordReset] Invalid code or expired. Now={timezone.now()}, Expires={getattr(pr, 'expires_at', None)}")
            form.add_error('code', 'Invalid or expired code. Please request a new password reset.')

    else:
        # Prefill form with GET params (for auto entry from email link)
        initial = {}
        code_param = request.GET.get('code') or request.GET.get('c')
        id_param = request.GET.get('id') or request.GET.get('identifier')
        if code_param:
            initial['code'] = code_param.strip()
        if id_param:
            initial['email_or_index'] = id_param.strip()
        form = PasswordResetConfirmForm(initial=initial)

    return render(request, 'registration/password_reset_confirm.html', {'form': form})

def login_view(request):
    message = None

    if request.method == "POST":
        credential = request.POST.get("credential", "").strip()
        password = request.POST.get("password", "").strip()
        user = None

        if "@" in credential:
            user = User.objects.filter(email__iexact=credential).first()
        else:
            profile = (
                StudentProfile.objects.filter(index_number__iexact=credential).first()
                or StudentProfile.objects.filter(phone_number__iexact=credential).first()
                or StaffProfile.objects.filter(staff_id__iexact=credential).first()
            )
            if profile:
                user = profile.user

        if user and user.check_password(password):
            login(request, user)

            subject = "Welcome Back to the Transcript Portal"
            message_body = (
                f"Hello {user.username},\n\n"
                f"Welcome back! You successfully logged into the Transcript Request Portal on "
                f"{timezone.now().strftime('%Y-%m-%d %H:%M:%S')}.\n\n"
                "You can now manage your transcript requests or check your dashboard for updates.\n\n"
                "Best regards,\n"
                "Academic Records Office"
            )

            try:
                send_mail(
                    subject,
                    message_body,
                    settings.EMAIL_HOST_USER,
                    [user.email],
                    fail_silently=True,
                )
            except Exception as e:
                print(f"Email sending failed: {e}")

            return redirect("staff_dashboard" if user.is_staff else "student_dashboard")

        message = "Invalid credentials"

    return render(request, "login.html", {"message": message})



@login_required
def logout_view(request):
    logout(request)
    return redirect("login")


def profile_pic(request):
    return render(request, 'profile_pic.html')

@staff_member_required  
def create_staff(request):
    if not request.user.is_superuser:
        messages.error(request, "Only superusers can create staff accounts.")
        return redirect('staff_dashboard')
    if request.method == "POST":
        form = StaffSignupForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Staff account created successfully.")
            return redirect("login")
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = StaffSignupForm()
    return render(request, "staff/create_staff.html", {"form": form})




from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.db.models import OuterRef, Subquery, Prefetch
from django.utils import timezone

from .models import (
    StudentProfile,
    Student,
    TranscriptRequest,
    Transcript,
    TranscriptStatus,
)


from django.shortcuts import render, redirect
from django.db.models import Prefetch, OuterRef, Subquery
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from .models import StudentProfile, Student, TranscriptRequest, TranscriptStatus, Transcript

@login_required
def student_dashboard(request):
    # Redirect staff users to staff dashboard
    if request.user.is_staff:
        return redirect('staff_dashboard')

    # Try to get student's profile and related info
    profile = StudentProfile.objects.select_related(
        'student', 'student__program', 'student__department'
    ).filter(user=request.user).first()

    student = None
    requests_qs = TranscriptRequest.objects.none()

    if profile:
        student = profile.student

        # Attempt to auto-link student record by index number if missing
        if not student and profile.index_number:
            linked_student = Student.objects.select_related('program', 'department').filter(
                index_number=profile.index_number
            ).first()
            if linked_student:
                profile.student = linked_student
                profile.save(update_fields=['student'])
                student = linked_student

        # If student exists, fetch their transcript requests
        if student:
            requests_qs = (
                TranscriptRequest.objects.filter(student=student)
                .order_by('-date_requested')
                .prefetch_related(
                    'statuses',
                    'selections__batch',
                    Prefetch('transcript', queryset=Transcript.objects.filter(file__isnull=False)),
                    'fee_clearance',
                )
            )

    # ----- BUILD REQUEST DATA -----
    requests_data = []
    for req in requests_qs:
        latest_status = req.statuses.order_by('-updated_on', '-id').first()
        fc = getattr(req, 'fee_clearance', None)
        payment_made = getattr(req, 'payment_made', False)

        # ----- PAYMENT STATUS -----
        if not fc:
            payment_label = ("Awaiting Review", "info")
        else:
            owes = getattr(fc, 'owes', False)
            cleared = getattr(fc, 'cleared', False)
            amount_owed = getattr(fc, 'amount_owed', None)

            if owes and not payment_made:
                payment_label = (f"Outstanding: GHS {amount_owed or '—'}", "warning")
            elif owes and payment_made and not cleared:
                payment_label = ("Paid - Awaiting Verification", "info")
            elif payment_made and cleared:
                payment_label = ("Paid & Cleared", "success")
            else:
                payment_label = ("No Payment Required", "light")

        # ----- PROCESSING STATUS -----
        if latest_status:
            stage_map = {
                'pending': ("Awaiting Review", "info"),
                'accounts_office': ("At Accounts Office", "warning"),
                'exams_office': ("At Exams Office", "primary"),
                'registrar': ("At Registrar", "secondary"),
                'approved': ("Approved", "success"),
                'rejected': (f"Disapproved by {latest_status.updated_by}", "danger"),
            }
            processing_label = stage_map.get(
                latest_status.stage,
                ((latest_status.stage or "Processing").title(), "dark")
            )
        else:
            processing_label = ("Awaiting Processing", "secondary")

        # ----- WHO PROCESSED -----
        try:
            if latest_status and not hasattr(latest_status, 'processed_by'):
                pb = getattr(latest_status, 'updated_by', None)
                if not pb and getattr(latest_status, 'user', None):
                    pb = getattr(latest_status.user, 'get_full_name', lambda: None)() or getattr(latest_status.user, 'username', None)
                latest_status.processed_by = pb
        except Exception:
            pass

        # ----- TRANSCRIPT & DOWNLOAD LOGIC -----
        transcript_obj = getattr(req, 'transcript', None)
        has_file = transcript_obj and transcript_obj.file

        # ✅ Simplified: Allow download if transcript file exists
        if has_file:
            if req.transcript_type == 'unofficial':
                download_ready = True
            else:
                download_ready = latest_status and latest_status.stage in ['approved', 'registrar']
        else:
            download_ready = False

        # ----- REJECTION DETAILS -----
        rejection_details = None
        rejection_file = None
        if latest_status and latest_status.stage == "rejected":
            rejection_details = latest_status.remarks
            if hasattr(latest_status, 'attachment') and latest_status.attachment:
                rejection_file = latest_status.attachment.url

        # ----- BUILD ENTRY -----
        requests_data.append({
            'request': req,
            'latest_status': latest_status,
            'payment_label': payment_label[0],
            'payment_badge': payment_label[1],
            'processing_label': processing_label[0],
            'processing_badge': processing_label[1],
            'download_ready': download_ready,
            'transcript_file': transcript_obj.file.url if has_file else None,
            'rejection_details': rejection_details,
            'rejection_file': rejection_file,
        })

    # ----- APPROVED TRANSCRIPTS -----
    approved_transcripts = []
    new_approved = False
    if student:
        latest_status_subquery = TranscriptStatus.objects.filter(
            transcript_request=OuterRef('pk')
        ).order_by('-updated_on')

        approved_transcripts = Transcript.objects.select_related('transcript_request').filter(
            transcript_request__student=student,
            file__isnull=False,
            transcript_request__statuses__stage=Subquery(latest_status_subquery.values('stage')[:1]),
        ).filter(
            transcript_request__statuses__stage='approved'
        ).distinct().order_by('-date_generated')

        recent = timezone.now() - timezone.timedelta(hours=24)
        new_approved = approved_transcripts.filter(date_generated__gte=recent).exists()

    context = {
        'profile': profile,
        'student': student,
        'requests_data': requests_data,
        'approved_transcripts': approved_transcripts,
        'new_approved': new_approved,
    }

    return render(request, "student_dashboard.html", context)



from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils import timezone
from .models import TranscriptRequest


def export_staff_payments_excel(request):
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    payments = TranscriptRequest.objects.filter(payment_made=True).select_related("student", "student__user")

    if start_date:
        payments = payments.filter(date_requested__date__gte=start_date)
    if end_date:
        payments = payments.filter(date_requested__date__lte=end_date)

    grouped = {}
    for p in payments:
        date_str = p.date_requested.strftime("%Y-%m-%d")
        grouped.setdefault(date_str, []).append(p)

    wb = Workbook()
    ws = wb.active
    ws.title = "Transcript Payment Report"

    # === Color & Style Setup ===
    dark_blue = "1F4E78"
    sky_blue = "D9EAF7"
    silver = "E9ECEF"
    white = "FFFFFF"
    accent = "007BFF"
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # === Title Section ===
    ws.merge_cells("A1:E1")
    ws["A1"] = "Transcript Payment Summary"
    ws["A1"].font = Font(size=15, bold=True, color=white)
    ws["A1"].fill = PatternFill(start_color=dark_blue, end_color=dark_blue, fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:E2")
    ws["A2"] = f"Generated on {timezone.now().strftime('%B %d, %Y - %I:%M %p')}"
    ws["A2"].font = Font(size=10, color="EEEEEE", italic=True)
    ws["A2"].fill = PatternFill(start_color=dark_blue, end_color=dark_blue, fill_type="solid")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 20

    row_num = 4
    grand_total = 0

    # === Loop over each day ===
    for i, (date, day_payments) in enumerate(grouped.items(), start=1):
        # Date section header
        ws.merge_cells(f"A{row_num}:E{row_num}")
        ws[f"A{row_num}"] = f"📅 {timezone.datetime.strptime(date, '%Y-%m-%d').strftime('%A, %B %d, %Y')}"
        ws[f"A{row_num}"].font = Font(bold=True, color=white)
        ws[f"A{row_num}"].fill = PatternFill(start_color=accent, end_color=accent, fill_type="solid")
        ws[f"A{row_num}"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row_num].height = 22
        row_num += 1

        # Table header
        headers = ["#", "Student", "Transcript Type", "Date Requested", "Amount (GHS)"]
        for col_num, title in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num, value=title)
            cell.font = Font(bold=True, color=dark_blue)
            cell.fill = PatternFill(start_color=silver, end_color=silver, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        row_num += 1

        # Table rows
        day_total = 0
        for idx, p in enumerate(day_payments, start=1):
            bg = white if idx % 2 else sky_blue
            student_name = (
                p.student.user.get_full_name()
                if hasattr(p.student, "user") and p.student.user
                else str(p.student)
            )
            row = [
                idx,
                student_name,
                p.transcript_type or "—",
                p.date_requested.strftime("%b %d, %Y"),
                float(p.amount or 0),
            ]
            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
                cell.border = border
                cell.alignment = Alignment(
                    horizontal="right" if col_num == 5 else "left", vertical="center"
                )
                if col_num == 1:
                    cell.alignment = Alignment(horizontal="center")
            day_total += float(p.amount or 0)
            row_num += 1

        # Day total row
        ws.merge_cells(f"A{row_num}:D{row_num}")
        ws[f"A{row_num}"] = "Subtotal"
        ws[f"A{row_num}"].font = Font(bold=True, color=dark_blue)
        ws[f"A{row_num}"].alignment = Alignment(horizontal="right")
        ws[f"E{row_num}"] = day_total
        ws[f"E{row_num}"].font = Font(bold=True, color=accent)
        ws[f"E{row_num}"].alignment = Alignment(horizontal="right")
        ws[f"E{row_num}"].border = border
        row_num += 2
        grand_total += day_total

    # === Grand Total Summary ===
    ws.merge_cells(f"A{row_num}:D{row_num}")
    ws[f"A{row_num}"] = "Grand Total"
    ws[f"A{row_num}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"A{row_num}"].fill = PatternFill(start_color=dark_blue, end_color=dark_blue, fill_type="solid")
    ws[f"A{row_num}"].alignment = Alignment(horizontal="right", vertical="center")

    ws[f"E{row_num}"] = grand_total
    ws[f"E{row_num}"].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f"E{row_num}"].fill = PatternFill(start_color=dark_blue, end_color=dark_blue, fill_type="solid")
    ws[f"E{row_num}"].alignment = Alignment(horizontal="right", vertical="center")

    ws.row_dimensions[row_num].height = 24

    # === Column width auto-adjust ===
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 25

    # === Response ===
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"Staff_Transcript_Payments_{timezone.now().strftime('%Y%m%d')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from .models import Payment


@login_required
def registrar_export_payments(request):
    # Fetch and sort payments by date
    payments = (
        Payment.objects.select_related("transcript_request__student")
        .order_by("date_checked")
    )

    # Group payments by date
    grouped = {}
    for p in payments:
        date_key = p.date_checked.strftime("%Y-%m-%d") if p.date_checked else "Unknown"
        grouped.setdefault(date_key, []).append(p)

    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Registrar Payments"

    # === STYLE SETUP ===
    dark_blue = "1F4E78"
    soft_gray = "F2F2F2"
    light_blue = "E9F3FB"
    total_blue = "007BFF"
    white = "FFFFFF"

    border_side = Side(style="thin", color="D9D9D9")
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    currency_style = NamedStyle(name="currency_style")
    currency_style.number_format = "₵#,##0.00"

    # === HEADER ===
    ws.merge_cells("A1:G1")
    ws["A1"] = "Registrar Transcript Payment Summary"
    ws["A1"].font = Font(bold=True, size=14, color=white)
    ws["A1"].fill = PatternFill(start_color=dark_blue, end_color=dark_blue, fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:G2")
    ws["A2"] = f"Generated on {timezone.now().strftime('%B %d, %Y — %I:%M %p')}"
    ws["A2"].font = Font(size=10, italic=True, color="E9ECEF")
    ws["A2"].fill = PatternFill(start_color=dark_blue, end_color=dark_blue, fill_type="solid")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    # === COLUMN HEADERS ===
    headers = [
        "Date",
        "Student Name",
        "Index Number",
        "Reference Code",
        "Amount (₵)",
        "Status",
        "Checked By",
    ]

    row_num = 4
    for col_num, title in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_num, value=title)
        cell.font = Font(bold=True, color=dark_blue)
        cell.fill = PatternFill(start_color=soft_gray, end_color=soft_gray, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    row_num += 1

    grand_total = 0

    # === DATA ROWS ===
    for date_key, records in grouped.items():
        # Date label row (section header)
        ws.merge_cells(f"A{row_num}:G{row_num}")
        ws[f"A{row_num}"] = (
            f"📅 {timezone.datetime.strptime(date_key, '%Y-%m-%d').strftime('%A, %B %d, %Y')}"
            if date_key != "Unknown"
            else "📅 Unknown Date"
        )
        ws[f"A{row_num}"].font = Font(bold=True, color=white)
        ws[f"A{row_num}"].fill = PatternFill(start_color=total_blue, end_color=total_blue, fill_type="solid")
        ws[f"A{row_num}"].alignment = Alignment(horizontal="left", vertical="center")
        row_num += 1

        day_total = 0
        for idx, pay in enumerate(records, start=1):
            student = pay.transcript_request.student
            row = [
                pay.date_checked.strftime("%Y-%m-%d") if pay.date_checked else "—",
                student.name,
                student.index_number,
                pay.transcript_request.reference_code,
                float(pay.transcript_request.amount or 0),
                "Cleared" if pay.cleared else "Pending",
                getattr(pay.checked_by, "username", "—"),
            ]

            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if col_num == 5:  # Amount column
                    cell.style = currency_style
                if idx % 2 == 0:
                    cell.fill = PatternFill(start_color=light_blue, end_color=light_blue, fill_type="solid")

            day_total += float(pay.transcript_request.amount or 0)
            row_num += 1

        # === SUBTOTAL ROW ===
        ws.merge_cells(f"A{row_num}:D{row_num}")
        ws[f"A{row_num}"] = f"Subtotal ({date_key})"
        ws[f"A{row_num}"].font = Font(bold=True, color=dark_blue)
        ws[f"A{row_num}"].alignment = Alignment(horizontal="right", vertical="center")
        ws[f"A{row_num}"].fill = PatternFill(start_color=soft_gray, end_color=soft_gray, fill_type="solid")
        ws[f"A{row_num}"].border = border

        ws[f"E{row_num}"] = day_total
        ws[f"E{row_num}"].style = currency_style
        ws[f"E{row_num}"].font = Font(bold=True, color=dark_blue)
        ws[f"E{row_num}"].fill = PatternFill(start_color=soft_gray, end_color=soft_gray, fill_type="solid")
        ws[f"E{row_num}"].alignment = Alignment(horizontal="right", vertical="center")
        for col in ["F", "G"]:
            ws[f"{col}{row_num}"].fill = PatternFill(start_color=soft_gray, end_color=soft_gray, fill_type="solid")
            ws[f"{col}{row_num}"].border = border

        grand_total += day_total
        row_num += 2

    # === GRAND TOTAL ROW ===
    ws.merge_cells(f"A{row_num}:D{row_num}")
    ws[f"A{row_num}"] = "GRAND TOTAL"
    ws[f"A{row_num}"].font = Font(bold=True, color=white)
    ws[f"A{row_num}"].fill = PatternFill(start_color=dark_blue, end_color=dark_blue, fill_type="solid")
    ws[f"A{row_num}"].alignment = Alignment(horizontal="right", vertical="center")

    ws[f"E{row_num}"] = grand_total
    ws[f"E{row_num}"].style = currency_style
    ws[f"E{row_num}"].font = Font(bold=True, color=white)
    ws[f"E{row_num}"].fill = PatternFill(start_color=dark_blue, end_color=dark_blue, fill_type="solid")
    ws[f"E{row_num}"].alignment = Alignment(horizontal="right", vertical="center")

    for col in ["F", "G"]:
        ws[f"{col}{row_num}"].fill = PatternFill(start_color=dark_blue, end_color=dark_blue, fill_type="solid")
        ws[f"{col}{row_num}"].border = border

    # === COLUMN WIDTHS ===
    widths = [15, 28, 18, 25, 15, 15, 20]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # === RESPONSE ===
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"Registrar_Payments_{timezone.now().strftime('%Y%m%d')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response



from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.db.models import Q, Sum, Subquery, OuterRef
from django.db.models.functions import TruncDate
from .models import StaffProfile, TranscriptRequest, TranscriptStatus, Transcript


@login_required
@staff_member_required
def staff_dashboard(request):
    profile = StaffProfile.objects.filter(user=request.user).first()
    role = profile.role.lower() if profile and getattr(profile, 'role', None) else None

    accounts_pending = accounts_processed = accounts_rejected = accounts_forwarded = []
    exams_queue = exams_processed = exams_disapproved = []
    registrar_pending = registrar_disapproved = []
    official_transcripts = unofficial_transcripts = []
    grand_total = 0
    grouped_payments = {}
    payments = []

    # ============ ACCOUNTS OFFICE ============
    if role == 'accounts_office':
        # Pending verification
        accounts_pending = (
            TranscriptRequest.objects.select_related('student', 'fee_clearance')
            .filter(transcript__isnull=True)
            .filter(Q(fee_clearance__isnull=True) | Q(payment_made=True, fee_clearance__cleared=False))
            .exclude(statuses__stage='exams_office')
            .order_by('-date_requested')
        )

        # Processed items
        accounts_processed = (
            TranscriptRequest.objects.select_related('student', 'fee_clearance')
            .filter(fee_clearance__cleared=True)
            .exclude(statuses__stage__in=['exams_office', 'registrar', 'approved', 'dispatched', 'rejected'])
            .distinct()
            .order_by('-date_requested')
        )

        # Rejected items
        accounts_rejected = (
            TranscriptRequest.objects.select_related('student', 'fee_clearance')
            .filter(fee_clearance__owes=True)
            .order_by('-date_requested')
        )

        # Forwarded to exams
        accounts_forwarded = (
            TranscriptRequest.objects.select_related('student', 'fee_clearance')
            .filter(Q(fee_clearance__cleared=True) | Q(statuses__stage='exams_office'))
            .distinct()
            .order_by('-date_requested')
        )

        # === Payments grouped by date ===
        daily_groups = (
            TranscriptRequest.objects.filter(payment_made=True)
            .annotate(payment_day=TruncDate('date_requested'))
            .values('payment_day')
            .annotate(total_amount=Sum('amount'))
            .order_by('-payment_day')
        )

        # Create dictionary: {date: {'payments': queryset, 'total': value}}
        grouped_payments = {}
        for day in daily_groups:
            date_key = day['payment_day']
            payments = (
                TranscriptRequest.objects.filter(payment_made=True, date_requested__date=date_key)
                .select_related('student')
                .order_by('-date_requested')
            )
            grouped_payments[date_key] = {
                'payments': payments,
                'total': day['total_amount']
            }

        # Grand total
        grand_total = sum(item['total_amount'] for item in daily_groups)

    # ============ EXAMS OFFICE ============
    elif role == 'exams_office':
        exams_queue = (
            TranscriptRequest.objects.select_related('student', 'fee_clearance')
            .filter(transcript__isnull=True)
            .filter(Q(statuses__stage='exams_office') | Q(payment_made=True, fee_clearance__cleared=True))
            .exclude(statuses__stage__in=['registrar', 'approved', 'dispatched', 'rejected'])
            .distinct()
            .order_by('-date_requested')
        )

        exams_processed = (
            TranscriptRequest.objects.select_related('student', 'fee_clearance', 'transcript')
            .filter(Q(statuses__stage__in=['registrar', 'approved', 'dispatched', 'rejected']) | Q(transcript__isnull=False))
            .distinct()
            .order_by('-date_requested')
        )

        # Only consider requests whose latest status is 'disapproved'.
        latest_status_for_req = TranscriptStatus.objects.filter(transcript_request=OuterRef('pk')).order_by('-updated_on')
        exams_disapproved = (
            TranscriptRequest.objects.select_related('student', 'fee_clearance', 'transcript')
            .filter(
                statuses__stage='disapproved',
                statuses__updated_on=Subquery(latest_status_for_req.values('updated_on')[:1])
            )
            .distinct()
            .order_by('-date_requested')
        )

        unofficial_transcripts = list(
            Transcript.objects.select_related('transcript_request__student')
            .filter(transcript_request__transcript_type='unofficial', file__isnull=False)
            .order_by('-date_generated')
        )

        official_transcripts = list(
            Transcript.objects.select_related('transcript_request__student')
            .filter(registrar_signature=True, transcript_request__transcript_type='official', file__isnull=False)
            .order_by('-date_generated')
        )

        # Add missing unofficial/official transcript requests not yet attached to transcript files
        included_unoff_ids = {t.transcript_request.id for t in unofficial_transcripts if getattr(t, 'transcript_request', None)}
        extra_unoff_qs = TranscriptRequest.objects.filter(
            transcript_type='unofficial', statuses__stage__in=['exams_office', 'approved']
        ).exclude(id__in=included_unoff_ids).distinct().order_by('-date_requested')
        for req in extra_unoff_qs:
            unofficial_transcripts.append(req)

        included_off_ids = {t.transcript_request.id for t in official_transcripts if getattr(t, 'transcript_request', None)}
        extra_off_qs = TranscriptRequest.objects.filter(
            transcript_type='official', statuses__stage__in=['registrar', 'approved']
        ).exclude(id__in=included_off_ids).distinct().order_by('-date_requested')
        for req in extra_off_qs:
            official_transcripts.append(req)

    # ============ REGISTRAR ============
    elif role == 'registrar':
        latest_status = TranscriptStatus.objects.filter(transcript_request=OuterRef('pk')).order_by('-updated_on')

        registrar_pending = (
            TranscriptRequest.objects.select_related('student', 'fee_clearance', 'transcript')
            .filter(
                statuses__stage='registrar',
                statuses__updated_on=Subquery(latest_status.values('updated_on')[:1])
            )
            .order_by('-date_requested')
        )

        registrar_disapproved = (
            TranscriptRequest.objects.select_related('student', 'fee_clearance', 'transcript')
            .filter(
                statuses__stage='rejected',
                statuses__updated_on=Subquery(latest_status.values('updated_on')[:1])
            )
            .order_by('-date_requested')
        )

        # === PAYMENTS TAB FILTER ===
        payments_qs = TranscriptRequest.objects.select_related('student').filter(payment_made=True)
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')

        if start_date:
            payments_qs = payments_qs.filter(date_requested__date__gte=start_date)
        if end_date:
            payments_qs = payments_qs.filter(date_requested__date__lte=end_date)

        payments = payments_qs.order_by('-date_requested')
        grand_total = payments_qs.aggregate(total=Sum('amount'))['total'] or 0

        # === TRANSCRIPTS ===
        official_transcripts = Transcript.objects.select_related('transcript_request__student').filter(
            transcript_request__transcript_type='official'
        ).filter(
            Q(registrar_signature=True) | Q(transcript_request__statuses__stage='approved')
        ).filter(file__isnull=False).distinct().order_by('-date_generated')

        unofficial_qs = Transcript.objects.select_related('transcript_request__student').filter(
            transcript_request__transcript_type='unofficial', file__isnull=False
        ).order_by('-date_generated')

        try:
            official_ids = list(
                official_transcripts.values_list('transcript_request_id', flat=True)
            )
            unofficial_transcripts = unofficial_qs.exclude(transcript_request__id__in=official_ids)
        except Exception:
            unofficial_transcripts = unofficial_qs

    # ==== FINAL CONTEXT ====
    def _ensure_pdf_url(item):
        t_obj = None
        if hasattr(item, 'file') and getattr(item, 'file', None):
            t_obj = item
        else:
            try:
                t_obj = getattr(item, 'transcript', None)
            except Exception:
                t_obj = None

        pdf_url = None
        try:
            if t_obj and getattr(t_obj, 'file', None):
                pdf_url = request.build_absolute_uri(t_obj.file.url)
        except Exception:
            pdf_url = None

        display_date = None
        try:
            if t_obj and getattr(t_obj, 'date_generated', None):
                display_date = t_obj.date_generated
            else:
                if hasattr(item, 'date_requested') and getattr(item, 'date_requested', None):
                    display_date = item.date_requested
                elif getattr(item, 'transcript_request', None) and getattr(item.transcript_request, 'date_requested', None):
                    display_date = item.transcript_request.date_requested
        except Exception:
            display_date = None

        return {
            'item': item,
            'pdf_url': pdf_url,
            'transcript_obj': t_obj,
            'display_date': display_date,
        }

    official_transcripts_norm = [_ensure_pdf_url(i) for i in official_transcripts] if official_transcripts else []
    unofficial_transcripts_norm = [_ensure_pdf_url(i) for i in unofficial_transcripts] if unofficial_transcripts else []

    context = {
        'staff_role': role,
        'accounts_pending': accounts_pending,
        'accounts_processed': accounts_processed,
        'accounts_rejected': accounts_rejected,
        'accounts_forwarded': accounts_forwarded,
        'exams_queue': exams_queue,
        'exams_processed': exams_processed,
        'exams_disapproved': exams_disapproved,
        'registrar_pending': registrar_pending,
        'registrar_disapproved': registrar_disapproved,
        'official_transcripts': official_transcripts_norm,
        'unofficial_transcripts': unofficial_transcripts_norm,
        'grouped_payments': grouped_payments,
        'grand_total': grand_total,
        'payments': payments,  # ✅ Now available for template
    }

    return render(request, "staff/staff_dashboard.html", context)


import io
from datetime import datetime, date
from django.http import HttpResponse
from django.db.models import Sum
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.db.models.functions import TruncDate
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from .models import TranscriptRequest


# ================= EXCEL EXPORT =================
@login_required
@staff_member_required
def export_payments_excel(request):
    """
    Export transcript payments to Excel, grouped by date,
    filtered by optional start_date and end_date.
    """
    # Get date filters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    qs = TranscriptRequest.objects.filter(payment_made=True)

    if start_date:
        qs = qs.filter(date_requested__date__gte=start_date)
    if end_date:
        qs = qs.filter(date_requested__date__lte=end_date)

    qs = qs.select_related('student').order_by('date_requested')

    # Group by date
    daily_groups = (
        qs.annotate(day=TruncDate('date_requested'))
        .values('day')
        .annotate(total_amount=Sum('amount'))
        .order_by('day')
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Transcript Payments"

    ws.append(["Transcript Payments Report"])
    ws.append([f"Period: {start_date or 'Beginning'} - {end_date or 'Today'}"])
    ws.append([])

    # Headers
    headers = ["Date", "Index", "Student Name", "Transcript Type", "Reference", "Amount (GHS)"]
    ws.append(headers)

    grand_total = 0.0

    for group in daily_groups:
        current_day = group['day']
        payments = qs.filter(date_requested__date=current_day)
        ws.append([f"{current_day}", "", "", "", "", ""])  # Date separator row

        daily_total = 0
        for req in payments:
            idx = req.student.index_number if req.student else ""
            name = req.student.name if req.student else ""
            ttype = req.get_transcript_type_display() if hasattr(req, 'get_transcript_type_display') else (req.transcript_type or "")
            ref = req.payment_reference or req.reference_code or ""
            amt = float(req.amount or 0)
            ws.append(["", idx, name, ttype, ref, amt])
            daily_total += amt

        ws.append(["", "", "", "", "Daily Total", daily_total])
        ws.append([])
        grand_total += daily_total

    ws.append([])
    ws.append(["", "", "", "", "GRAND TOTAL", grand_total])

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    filename = f"transcript_payments_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ================= PDF EXPORT =================
@login_required
@staff_member_required
def export_payments_pdf(request):
    """
    Export transcript payments to PDF, grouped by date,
    filtered by optional start_date and end_date.
    """
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    qs = TranscriptRequest.objects.filter(payment_made=True)
    if start_date:
        qs = qs.filter(date_requested__date__gte=start_date)
    if end_date:
        qs = qs.filter(date_requested__date__lte=end_date)

    qs = qs.select_related('student').order_by('date_requested')

    daily_groups = (
        qs.annotate(day=TruncDate('date_requested'))
        .values('day')
        .annotate(total_amount=Sum('amount'))
        .order_by('day')
    )

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    title = Paragraph("Transcript Payments Report", styles['Title'])
    period = Paragraph(f"Period: {start_date or 'Beginning'} - {end_date or 'Today'}", styles['Normal'])
    elements.extend([title, Spacer(1, 6), period, Spacer(1, 12)])

    grand_total = 0

    for group in daily_groups:
        day = group['day']
        payments = qs.filter(date_requested__date=day)
        elements.append(Paragraph(f"<b>Date: {day}</b>", styles['Heading4']))
        elements.append(Spacer(1, 6))

        data = [["#", "Index", "Student Name", "Type", "Reference", "Amount (GHS)"]]
        daily_total = 0

        for i, req in enumerate(payments, 1):
            idx = req.student.index_number if req.student else ""
            name = req.student.name if req.student else ""
            ttype = req.get_transcript_type_display() if hasattr(req, 'get_transcript_type_display') else (req.transcript_type or "")
            ref = req.payment_reference or req.reference_code or ""
            amt = float(req.amount or 0)
            data.append([i, idx, name, ttype, ref, f"{amt:,.2f}"])
            daily_total += amt

        data.append(["", "", "", "", "Daily Total", f"GHS {daily_total:,.2f}"])
        grand_total += daily_total

        table = Table(data, colWidths=[25, 70, 130, 80, 80, 70])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003366')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
            ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
        ]))
        elements.extend([table, Spacer(1, 12)])

    elements.append(Paragraph(f"<b>Grand Total: GHS {grand_total:,.2f}</b>", styles['Heading3']))

    doc.build(elements)
    buffer.seek(0)

    filename = f"transcript_payments_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response






@login_required
@staff_member_required
def registrar_reapprove(request, request_id):
    transcript_request = get_object_or_404(TranscriptRequest, id=request_id)

    TranscriptStatus.objects.create(
        transcript_request=transcript_request,
        stage='approved',
        remarks='Re-approved by Registrar after disapproval review.',
        updated_by=request.user.get_full_name() or request.user.username
    )

    messages.success(request, f"Transcript request for {transcript_request.student.name} has been re-approved.")
    return redirect('staff_dashboard')


from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.core.mail import send_mail
from django.conf import settings
from .models import (
    TranscriptRequest, StaffProfile, TranscriptStatus,
    FacultyRegistrar, StudentProfile
)
# from .utils import email_student
from django.db.models import Q

@login_required
def generate_transcript(request, pk):
    transcript_request = get_object_or_404(TranscriptRequest, pk=pk)
    existing_transcript = getattr(transcript_request, 'transcript', None)

    # ⚡ Get related student and profile info
    student = transcript_request.student
    student_profile = StudentProfile.objects.filter(student=student).select_related('student', 'user').first()
    program = getattr(student, 'program', None)
    department = getattr(program, 'department', None)

    if request.method == 'POST':
        staff = StaffProfile.objects.filter(user=request.user, role='exams_office').first()
        if not staff:
            messages.error(request, "Only Exams Office staff can process transcripts.")
            return redirect('staff_dashboard')

        clearance = getattr(transcript_request, 'fee_clearance', None)
        if transcript_request.transcript_type == 'official' and (not clearance or not clearance.cleared):
            messages.error(request, "Accounts verification required before transcript processing.")
            return redirect('staff_dashboard')

        action = request.POST.get('action')

        if transcript_request.transcript_type == 'unofficial':
            registrar_id = request.POST.get('registrar_id') or request.POST.get('registrar')
            if not registrar_id:
                messages.error(request, 'Please select the correct Faculty Registrar before forwarding.')
                return redirect('generate_transcript', pk=transcript_request.id)

            try:
                registrar_obj = FacultyRegistrar.objects.filter(id=int(registrar_id)).first()
            except Exception:
                registrar_obj = None

            if registrar_obj:
                transcript_request.selected_faculty_registrar = registrar_obj
                transcript_request.save()

            TranscriptStatus.objects.create(
                transcript_request=transcript_request,
                stage='registrar',
                updated_by=staff.user.get_full_name() or staff.user.username,
                remarks=f"Unofficial transcript prepared by Exams Office. Faculty Registrar used: {registrar_obj.name if registrar_obj else '(none)'}"
            )

            try:
                subject = "Unofficial Transcript Forwarded to Registrar"
                message_body = (
                    f"Dear {student.name},\n\n"
                    f"Your unofficial transcript has been reviewed by the Exams Office and "
                    f"forwarded to the Registrar for final approval.\n\n"
                    f"Reference Code: {transcript_request.reference_code}\n\n"
                    f"You will be notified once it is approved.\n\n"
                    f"Best regards,\nAcademic Records Office"
                )
                email_student(subject, message_body, student)
            except Exception as e:
                print(f"Email to student failed: {e}")

            try:
                registrar_staff = StaffProfile.objects.filter(role='registrar').select_related('user')
                registrar_emails = [s.user.email for s in registrar_staff if s.user.email]
                if registrar_emails:
                    subject = f"Unofficial Transcript Pending Approval - {transcript_request.reference_code}"
                    message = (
                        f"Dear Registrar,\n\n"
                        f"An unofficial transcript is pending your approval.\n\n"
                        f"Student: {student.name}\n"
                        f"Index Number: {student.index_number}\n"
                        f"Program: {student.program.name}\n"
                        f"Faculty Registrar: {registrar_obj.name if registrar_obj else 'N/A'}\n"
                        f"Reference Code: {transcript_request.reference_code}\n"
                        f"Prepared By: {staff.user.get_full_name() or staff.user.username}\n\n"
                        f"Please log into the system to review and approve the transcript.\n\n"
                        f"Best regards,\nExams Office"
                    )
                    send_mail(
                        subject,
                        message,
                        settings.DEFAULT_FROM_EMAIL,
                        registrar_emails,
                        fail_silently=True,
                    )
            except Exception as e:
                print(f"Failed to notify registrar: {e}")

            messages.success(request, "Unofficial transcript forwarded to Registrar for approval.")
            return redirect('request_detail', pk=transcript_request.id)

        elif transcript_request.transcript_type == 'official':
            TranscriptStatus.objects.create(
                transcript_request=transcript_request,
                stage='registrar',
                updated_by=staff.user.get_full_name() or staff.user.username,
                remarks="Official transcript prepared by Exams Office. Awaiting Registrar approval."
            )

            try:
                subject = "Official Transcript Forwarded to Registrar"
                message_body = (
                    f"Dear {student.name},\n\n"
                    f"Your official transcript has been reviewed by the Exams Office and "
                    f"forwarded to the Registrar for final approval.\n\n"
                    f"Reference Code: {transcript_request.reference_code}\n\n"
                    f"You will be notified once it is approved.\n\n"
                    f"Best regards,\nAcademic Records Office"
                )
                email_student(subject, message_body, student)
            except Exception as e:
                print(f"Email to student failed: {e}")

            # Notify registrar
            try:
                registrar_staff = StaffProfile.objects.filter(role='registrar').select_related('user')
                registrar_emails = [s.user.email for s in registrar_staff if s.user.email]
                if registrar_emails:
                    subject = f"Official Transcript Pending Approval - {transcript_request.reference_code}"
                    message = (
                        f"Dear Registrar,\n\n"
                        f"An official transcript is pending your approval.\n\n"
                        f"Student: {student.name}\n"
                        f"Index Number: {student.index_number}\n"
                        f"Program: {student.program.name}\n"
                        f"Reference Code: {transcript_request.reference_code}\n"
                        f"Prepared By: {staff.user.get_full_name() or staff.user.username}\n\n"
                        f"Please log into the system to review and approve the transcript.\n\n"
                        f"Best regards,\nExams Office"
                    )
                    send_mail(
                        subject,
                        message,
                        settings.DEFAULT_FROM_EMAIL,
                        registrar_emails,
                        fail_silently=True,
                    )
            except Exception as e:
                print(f"Failed to notify registrar: {e}")

            messages.success(request, "Official transcript forwarded to Registrar for approval.")
            return redirect('request_detail', pk=transcript_request.id)

    # Load available registrars + past selection
    registrars = FacultyRegistrar.objects.all()
    parsed_registrar_id = None
    try:
        st = transcript_request.statuses.filter(stage='exams_office').first()
        if st and st.remarks:
            import re
            m = re.search(r'Faculty Registrar used:\s*(.+)', st.remarks)
            if m:
                reg_name = m.group(1).strip()
                r = FacultyRegistrar.objects.filter(name__iexact=reg_name).first()
                if r:
                    parsed_registrar_id = r.id
    except Exception:
        parsed_registrar_id = None

    return render(request, "generate_transcript.html", {
        'transcript_request': transcript_request,
        'registrars': registrars,
        'parsed_registrar_id': parsed_registrar_id,
        # ⚡ Pass new context variables
        'student': student,
        'student_profile': student_profile,
        'program': program,
        'department': department,
    })


@login_required
def student_download_transcript(request, pk):
    transcript_request = get_object_or_404(TranscriptRequest, pk=pk)
    student_profile = StudentProfile.objects.filter(user=request.user).first()
    if student_profile and student_profile.student == transcript_request.student:
        pass  
    else:
        staff_profile = StaffProfile.objects.filter(user=request.user).first()
        if not staff_profile:
            messages.error(request, "You are not authorized to access this transcript.")
            return redirect('student_dashboard')

    transcript_type = transcript_request.transcript_type.lower()

    if transcript_type == 'official':
        is_approved = transcript_request.statuses.filter(stage__in=['registrar', 'approved']).exists()
    else:
        is_approved = transcript_request.statuses.filter(stage__in=['exams_office', 'approved']).exists()

    if not is_approved:
        messages.error(request, "This transcript is not approved yet.")
        return redirect('student_dashboard')

    transcript = Transcript.objects.filter(transcript_request=transcript_request, file__isnull=False).first()
    if transcript and transcript.file:
        # In development (DEBUG) or when explicitly requested via ?regen=1, regenerate the PDF
        from django.conf import settings
        regen = request.GET.get('regen') == '1'
        if not (getattr(settings, 'DEBUG', False) or regen):
            try:
                file_path = transcript.file.path
                filename = f"{transcript_request.reference_code}_{transcript_request.get_transcript_type_display()}.pdf"
                return FileResponse(open(file_path, 'rb'), as_attachment=True, filename=filename)
            except Exception:
                messages.error(request, "Unable to retrieve transcript file.")
                return redirect('student_dashboard')
    try:
        # Regenerate the PDF to ensure latest signatures are embedded before serving
        from .utils import generate_unofficial_transcript_pdf, generate_official_transcript_pdf
        base_url = request.build_absolute_uri('/')[:-1]

        if transcript_request.transcript_type == 'unofficial':
            faculty_registrar = transcript_request.selected_faculty_registrar
            transcript_obj, pdf_bytes = generate_unofficial_transcript_pdf(
                transcript_request, faculty_registrar, base_url
            )
        else:
            transcript_obj, pdf_bytes = generate_official_transcript_pdf(
                transcript_request, include_registrar=True, include_vc=True, base_url=base_url
            )

        # If a Transcript model instance exists, ensure flags are set and file saved
        if transcript_obj:
            try:
                # ensure boolean flags reflect presence of signatures
                transcript_obj.registrar_signature = bool(getattr(transcript_obj, 'registrar_signature', False))
                transcript_obj.vc_signature = bool(getattr(transcript_obj, 'vc_signature', False))
                transcript_obj.save()
            except Exception:
                pass

        filename = f"{transcript_request.reference_code}_{transcript_request.get_transcript_type_display()}.pdf"
        return FileResponse(BytesIO(pdf_bytes), as_attachment=True, filename=filename)

    except Exception as e:
        messages.error(request, f"Unable to generate transcript: {str(e)}")
        return redirect('student_dashboard')


import urllib.parse
import re
import os
from django.conf import settings
from django.http import FileResponse, Http404
from django.shortcuts import render
from .models import TranscriptRequest, Transcript


def serve_transcript_pdf(request, filename):
    """Serve the transcript PDF and allow embedding in iframe."""
    file_path = os.path.join(settings.MEDIA_ROOT, 'transcripts', filename)
    if not os.path.exists(file_path):
        raise Http404("Transcript not found")

    response = FileResponse(open(file_path, 'rb'), content_type='application/pdf')
    response["X-Frame-Options"] = "ALLOWALL"
    return response


def verify_transcript(request, code):
    code_decoded = urllib.parse.unquote(code).strip()

    def clean_ref(r):
        """Remove 'ref', 'ref-', 'ref_' prefix if present."""
        return re.sub(r'(?i)^ref[-_]?','', (r or '') )

    # Try to find the transcript request
    tr = TranscriptRequest.objects.select_related(
        'student', 'student__program', 'student__department'
    ).filter(reference_code__iexact=code_decoded).first()

    if not tr:
        for i in range(4, 11):
            idx_try = code_decoded[:i]
            ref_try = clean_ref(code_decoded[i:])
            tr = TranscriptRequest.objects.select_related(
                'student', 'student__program', 'student__department'
            ).filter(reference_code__icontains=ref_try, student__index_number=idx_try).first()
            if tr:
                break

    if not tr:
        return render(request, 'verify_result.html', {
            'valid': False,
            'message': 'No matching transcript request found for the provided code.'
        })

    # Get transcript and faculty registrar safely
    transcript = getattr(tr, 'transcript', None)
    faculty_registrar = getattr(tr, 'selected_faculty_registrar', None)

    # Determine approved/authentic state. Prefer explicit verification/approval records
    verified_flag = False
    approved_flag = False
    try:
        if transcript:
            verified_flag = bool(getattr(getattr(transcript, 'verification', None), 'verified', False))
            approved_flag = bool(getattr(getattr(transcript, 'approval', None), 'approved', False))
    except Exception:
        verified_flag = False
        approved_flag = False

    # Fallback to status stages if no explicit verification/approval
    approved = verified_flag or approved_flag or tr.statuses.filter(stage='approved').exists()

    # Build the PDF URL for iframe preview
    if transcript and transcript.file:
        filename = transcript.file.name.split('/')[-1]
        pdf_url = request.build_absolute_uri(f"/view_transcript/{filename}/")
    else:
        pdf_url = None

    # Build a helpful message depending on what made it "approved"
    if verified_flag:
        message = 'Transcript is authentic (verified).'
    elif approved_flag:
        message = 'Transcript is authentic (approved).'
    elif approved:
        message = 'Record found and marked approved in status history.'
    else:
        message = 'Record found but not approved/dispatched yet.'

    context = {
        'valid': approved,
        'transcript_request': tr,
        'transcript': transcript,
        'faculty_registrar': faculty_registrar,
        'message': message,
        'pdf_url': pdf_url,
        'features_img': "https://fcache1.pakwheels.com/original/3X/d/1/d1c233aa776fa10db2af6d076c31f46709c41aeb.jpg",
    }

    return render(request, 'verify_result.html', context)




from django.shortcuts import render, redirect
from django.urls import reverse
from django.contrib import messages

def manual_verify(request):
    if request.method == 'POST':
        code = request.POST.get('reference_code', '').strip()

        if code:
            return redirect(reverse('verify_transcript', args=[code]))
        else:
            messages.error(request, "Please enter a valid reference number.")

    return render(request, 'manual_verify.html')




from django.http import FileResponse, Http404
from django.conf import settings
import os

def serve_transcript_pdf(request, filename):
    file_path = os.path.join(settings.MEDIA_ROOT, 'transcripts', filename)
    if not os.path.exists(file_path):
        raise Http404("Transcript not found")

    response = FileResponse(open(file_path, 'rb'), content_type='application/pdf')
    response["X-Frame-Options"] = "ALLOWALL"
    return response




def qr_scanner(request):
    """QR code scanner interface for transcript verification"""
    return render(request, 'qr_scanner.html')


@login_required
def download_invoice(request, pk):
    tr = get_object_or_404(TranscriptRequest, pk=pk)
    profile = StudentProfile.objects.select_related('student').filter(user=request.user).first()
    if not profile or profile.student != tr.student:
        messages.error(request, "You are not authorized to download this invoice.")
        return redirect('student_dashboard')

    clearance = getattr(tr, 'fee_clearance', None)
    if not clearance or not clearance.invoice_file:
        messages.error(request, "No invoice file attached for this request.")
        return redirect('student_dashboard')

    try:
        file_path = clearance.invoice_file.path
        return FileResponse(open(file_path, 'rb'), as_attachment=True, filename=file_path.split('/')[-1])
    except Exception as e:
        messages.error(request, "Unable to retrieve invoice file.")
        return redirect('student_dashboard')



def parse_date(date_value):
    if not date_value:
        return None
    try:
        if isinstance(date_value, datetime):
            return date_value.date()
        return datetime.strptime(str(date_value), "%Y-%m-%d").date()
    except ValueError:
        try:
            return datetime.strptime(str(date_value), "%m/%d/%Y").date()
        except ValueError:
            raise ValueError(f"Invalid date format: {date_value}. Expected YYYY-MM-DD or MM/DD/YYYY.")


@login_required
def upload_students(request):
    upload_form = StudentUploadForm()
    student_form = StudentForm()
    uploaded_count = 0
    errors = []

    if request.method == "POST":
        if "file_upload" in request.POST:
            upload_form = StudentUploadForm(request.POST, request.FILES)
            if upload_form.is_valid():
                file = request.FILES["file"]

                if file.name.endswith(".csv"):
                    data = TextIOWrapper(file.file, encoding="utf-8")
                    reader = csv.DictReader(data)
                    for i, row in enumerate(reader, start=2):
                        try:
                            index_number = str(row.get("index_number", "")).strip()
                            name = str(row.get("name", "")).strip()
                            program_name = str(row.get("program_name", "")).strip()
                            department_name = str(row.get("department_name", "")).strip()
                            owes_fees_raw = row.get("owes_fees", "")

                            if not index_number or not name:
                                raise ValueError("Missing required 'index_number' or 'name'.")

                            department, _ = Department.objects.get_or_create(
                                department=department_name or "General"
                            )
                            program, _ = Program.objects.get_or_create(
                                name=program_name or "General Program",
                                department=department
                            )

                            date_entered = parse_date(str(row.get("date_entered", "")).strip()) or None
                            date_completed = parse_date(str(row.get("date_completed", "")).strip()) or None

                            Student.objects.update_or_create(
                                index_number=index_number,
                                defaults={
                                    "name": name,
                                    "date_entered": date_entered,
                                    "date_completed": date_completed,
                                    "program": program,
                                    "department": department,
                                    "owes_fees": str(owes_fees_raw).strip().upper() in ["YES", "TRUE", "1"],
                                },
                            )
                            uploaded_count += 1

                        except Exception as e:
                            errors.append(f"Row {i}: {e}")

                elif file.name.endswith((".xlsx", ".xls")):
                    wb = openpyxl.load_workbook(file)
                    sheet = wb.active
                    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                        try:
                            name, index_number, date_entered_raw, date_completed_raw, program_name, department_name, owes_fees = row

                            index_number = str(index_number).strip()
                            name = str(name).strip()
                            program_name = str(program_name or "").strip()
                            department_name = str(department_name or "").strip()

                            if not index_number or not name:
                                raise ValueError("Missing required 'index_number' or 'name'.")

                            department, _ = Department.objects.get_or_create(
                                department=department_name or "General"
                            )
                            program, _ = Program.objects.get_or_create(
                                name=program_name or "General Program",
                                department=department
                            )

                            date_entered = parse_date(str(date_entered_raw).strip()) or None
                            date_completed = parse_date(str(date_completed_raw).strip()) or None

                        
                            Student.objects.update_or_create(
                                index_number=index_number,
                                defaults={
                                    "name": name,
                                    "date_entered": date_entered,
                                    "date_completed": date_completed,
                                    "program": program,
                                    "department": department,
                                    "owes_fees": str(owes_fees or "").strip().upper() in ["YES", "TRUE", "1"],
                                },
                            )
                            uploaded_count += 1

                        except Exception as e:
                            errors.append(f"Row {i}: {e}")
                else:
                    messages.error(request, "Please upload a valid CSV or Excel file.")
                    return redirect("upload_students")

                # Final messages
                if uploaded_count:
                    messages.success(request, f"{uploaded_count} students uploaded successfully!")
                if errors:
                    # keep HTML safe if your template renders messages as safe; otherwise consider storing in logs
                    messages.warning(request, f"Some rows failed:<br>{'<br>'.join(errors)}")
                return redirect("upload_students")

        # MANUAL ADD
        elif "manual_add" in request.POST:
            student_form = StudentForm(request.POST)
            if student_form.is_valid():
                student_form.save()
                messages.success(request, "Student added successfully!")
                return redirect("upload_students")
            else:
                messages.error(request, "Please correct the errors in the manual entry form.")

    context = {
        "upload_form": upload_form,
        "student_form": student_form,
    }
    return render(request, "upload_students.html", context)



def manage_students(request):
    return render(request, 'manage_students.html')


@csrf_exempt
def momo_callback(request):
    if request.method != 'POST':
        return HttpResponseNotFound('Not found')

    try:
        raw = request.body
        signature = request.headers.get('X-Signature', '')
        if not settings.DEBUG:
            secret = settings.MOMO_WEBHOOK_SECRET.encode('utf-8')
            expected = hmac.new(secret, raw, hashlib.sha256).hexdigest()
            if not hmac.compare_digest(signature, expected):
                return HttpResponseNotFound('Invalid signature')

        payload = json.loads(raw.decode('utf-8'))
        reference = payload.get('reference')
        original_reference = payload.get('original_reference') or reference
        status_ok = str(payload.get('status')).lower() == 'success'
        amount_paid = payload.get('amount')

        tr = TranscriptRequest.objects.select_related('student').filter(reference_code=original_reference).first()
        if not tr:
            return HttpResponseNotFound('Unknown reference')

        if status_ok:
            tr.payment_made = True
            tr.save(update_fields=['payment_made'])
            # Create or update a Payment record to persist the webhook information
            try:
                payment, created = Payment.objects.get_or_create(transcript_request=tr, defaults={'officer_name': 'System Auto'})
                # Attempt to coerce numeric amount if provided as string
                try:
                    payment.amount_received = float(amount_paid) if amount_paid is not None else payment.amount_received
                except Exception:
                    # leave as-is if conversion fails
                    pass
                # By default allocate received amount to accounts office until manual splitting is done
                if payment.amount_accounts_office is None:
                    payment.amount_accounts_office = payment.amount_received
                payment.cleared = True
                payment.notes = (payment.notes or '') + f"\nAuto webhook: Ref {reference} at {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
                payment.save()
            except Exception:
                # If saving payment fails, still continue to record status
                pass

            TranscriptStatus.objects.create(
                transcript_request=tr,
                stage='accounts_office',
                updated_by='System Auto',
                remarks=f"Payment received via Paystack. Amount: GHS {amount_paid}. Ref: {reference}. Time: {timezone.now().strftime('%Y-%m-%d %H:%M:%S') }"
            )
        return render(request, 'verify_result.html', {
            'valid': status_ok,
            'message': 'Payment processed successfully.' if status_ok else 'Payment failed.'
        })
    except Exception as e:
        return render(request, 'verify_result.html', {'valid': False, 'message': f'Error: {e}'})



@login_required
def student_approved_transcripts(request):
    profile = StudentProfile.objects.select_related('student').filter(user=request.user).first()
    if not profile or not profile.student:
        messages.error(request, "No student record found for your account.")
        return redirect('student_dashboard')

    student = profile.student
    approved_reqs_q = (
        Q(transcript__approval__approved=True)
        | Q(selections__batch__status='approved')
        | Q(statuses__stage='approved')
    )

    approved_requests = TranscriptRequest.objects.filter(
        student=student
    ).filter(approved_reqs_q).distinct().order_by('-date_requested')

    # Build a list of items (request + optional transcript + computed approved date)
    approved_items = []
    for req in approved_requests:
        transcript = getattr(req, 'transcript', None)
        # Prefer transcript.date_generated for approved date, fall back to the approved status timestamp
        approved_date = None
        if transcript and transcript.date_generated:
            approved_date = transcript.date_generated
        else:
            st = req.statuses.filter(stage='approved').first()
            approved_date = st.updated_on if st else None

        approved_items.append({
            'request': req,
            'transcript': transcript,
            'approved_date': approved_date,
        })

    return render(request, 'student_approved_transcripts.html', {
        'approved_transcripts': approved_items,
        'student': student,
    })

@login_required
def download_transcript(request, pk):
    """Allow student to download their approved transcript PDF"""
    transcript_request = get_object_or_404(
        TranscriptRequest,
        pk=pk,
        student__user=request.user
    )

    transcript = Transcript.objects.filter(transcript_request=transcript_request).first()
    if not transcript or not transcript.file:
        messages.error(request, "Transcript file not available for download.")
        return redirect('student_approved_transcripts')

    # Serve the PDF file
    file_path = transcript.file.path
    with open(file_path, 'rb') as f:
        response = HttpResponse(f.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{transcript.transcript_request.reference_code}.pdf"'
        return response

@login_required
def notify_payment_made(request, pk):
    """Student notifies accounts office that they have made a bank payment"""
    tr = get_object_or_404(TranscriptRequest, pk=pk)
    profile = StudentProfile.objects.select_related('student').filter(user=request.user).first()
    
    if not profile or profile.student != tr.student:
        messages.error(request, "You are not authorized for this action.")
        return redirect('student_dashboard')

    clearance = getattr(tr, 'fee_clearance', None)
    if not clearance or not clearance.owes:
        messages.error(request, "No outstanding payment found for this request.")
        return redirect('student_dashboard')

    if request.method == 'POST':
        tr.payment_made = True
        tr.save(update_fields=['payment_made'])

        TranscriptStatus.objects.create(
            transcript_request=tr,
            stage='accounts_office',
            updated_by=f"Student: {profile.student.name}",
            remarks=f"Student notified that bank payment has been made. Amount: GHS {clearance.amount_owed}. Awaiting accounts office verification."
        )

        try:
            from django.core.mail import send_mail
            from django.conf import settings
            
            accounts_staff = StaffProfile.objects.filter(role='accounts_office').select_related('user')
            accounts_emails = [staff.user.email for staff in accounts_staff if staff.user.email]
            
            if accounts_emails:
                subject = f"Payment Notification - {tr.reference_code}"
                message = (
                    f"Student Payment Notification\n\n"
                    f"Student: {profile.student.name}\n"
                    f"Index Number: {profile.student.index_number}\n"
                    f"Reference Code: {tr.reference_code}\n"
                    f"Amount Paid: GHS {clearance.amount_owed}\n"
                    f"Payment Method: Bank Transfer\n"
                    f"Notification Time: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
                    f"Please verify this payment and update the student's request status.\n\n"
                    f"Best regards,\n"
                    f"Transcript Management System"
                )
                
                send_mail(
                    subject,
                    message,
                    settings.DEFAULT_FROM_EMAIL,
                    accounts_emails,
                    fail_silently=True,
                )
        except Exception as e:
            print(f"Failed to send email notification to accounts office: {e}")

        messages.success(request, "Payment notification sent to Accounts Office. They will verify your payment and update your request status.")
        return redirect('student_dashboard')

    return render(request, 'notify_payment.html', {
        'transcript_request': tr,
        'amount_owed': clearance.amount_owed,
    })


@login_required
@staff_member_required
def accounts_verification_queue(request):
    if not request.user.is_staff:
        return redirect('login')
    staff = StaffProfile.objects.filter(user=request.user, role='accounts_office').first()
    if not staff:
        messages.error(request, 'Only Accounts Office staff can access this page.')
        return redirect('staff_dashboard')
    # Include requests where students have made payment but haven't been cleared yet
    items = TranscriptRequest.objects.filter(
        payment_made=True
    ).exclude(
        fee_clearance__cleared=True
    ).select_related('student').order_by('-date_requested')
    return render(request, 'accounts_verification_queue.html', { 'items': items })


@login_required
@staff_member_required
def accounts_verify_request(request, pk):
    if not request.user.is_staff:
        return redirect('login')
    staff = StaffProfile.objects.filter(user=request.user, role='accounts_office').first()
    if not staff:
        messages.error(request, 'Only Accounts Office staff can verify payments.')
        return redirect('staff_dashboard')
    tr = get_object_or_404(TranscriptRequest, pk=pk)
    FeeClearance.objects.update_or_create(
        transcript_request=tr,
        defaults={
            'cleared': True,
            'owes': False,
            'amount_owed': 0,
            'remarks': 'Cleared after payment verification by Accounts.',
            'officer_name': request.user.get_full_name() or request.user.username,
        }
    )
    TranscriptStatus.objects.create(
        transcript_request=tr,
        stage='accounts_office',
        updated_by=request.user.get_full_name() or request.user.username,
        remarks='Payment verified and cleared.'
    )
    TranscriptStatus.objects.create(
        transcript_request=tr,
        stage='exams_office',
        updated_by='System Auto',
        remarks='Forwarded to Exams Office for transcript generation.'
    )
    try:
        student = tr.student
        subject = "Transcript Payment Verified"
        message_body = (
            f"Dear {student.name},\n\n"
            f"Your payment has been verified by the Accounts Office. "
            f"Your dashboard has been updated and your request has been forwarded to the Exams Office for processing.\n\n"
            f"Reference: {tr.reference_code}\n\n"
            f"Best regards,\n"
            f"Academic Records Office"
        )

        # Prefer the email_student helper which resolves StudentProfile and user emails
        try:
            email_sent = False
            try:
                from .views import email_student
            except Exception:
                email_student = globals().get('email_student')

            if callable(email_student):
                try:
                    email_sent = bool(email_student(subject, message_body, student))
                except Exception:
                    email_sent = False

            # Fallback: resolve direct recipient and use safe_send_mail
            if not email_sent:
                recipient = None
                try:
                    recipient = getattr(getattr(student, 'user', None), 'email', None)
                except Exception:
                    recipient = None

                if not recipient:
                    try:
                        sp = StudentProfile.objects.select_related('user').filter(student=student).first()
                        if sp and getattr(sp.user, 'email', None):
                            recipient = sp.user.email
                    except Exception:
                        recipient = None

                if recipient:
                    safe_send_mail(subject, message_body, recipient)

        except Exception as e:
            print(f"Email notify (queue verify) failed: {e}")
    except Exception as e:
        print(f"Email notify (queue verify) outer failed: {e}")
    messages.success(request, 'Payment verified and forwarded to Exams Office.')
    return redirect('accounts_verification_queue')



from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib import messages
from .forms import UserUpdateForm, StudentProfileForm, StaffProfileForm
from .models import StudentProfile, StaffProfile

@login_required
def profile_settings(request):
    user = request.user
    is_student = hasattr(user, 'studentprofile')
    is_staff_member = hasattr(user, 'staffprofile')

    user_form = UserUpdateForm(instance=user)
    profile_form = None

    if request.method == 'POST':
        user_form = UserUpdateForm(request.POST, instance=user)

        if is_student:
            profile_form = StudentProfileForm(request.POST, request.FILES, instance=user.studentprofile)
        elif is_staff_member:
            profile_form = StaffProfileForm(request.POST, request.FILES, instance=user.staffprofile)

        if user_form.is_valid() and (not profile_form or profile_form.is_valid()):
            user_form.save()
            if profile_form:
                profile_form.save()
            messages.success(request, "✅ Profile updated successfully.")
            return redirect('profile_settings')
        else:
            messages.error(request, "⚠️ Please correct the errors below.")
    else:
        if is_student:
            profile_form = StudentProfileForm(instance=user.studentprofile)
        elif is_staff_member:
            profile_form = StaffProfileForm(instance=user.staffprofile)

    context = {
        'user_form': user_form,
        'profile_form': profile_form,
        'is_student': is_student,
        'is_staff_member': is_staff_member,
    }

    return render(request, 'profile_settings.html', context)





from django.shortcuts import render, redirect
from django.contrib.auth.decorators import user_passes_test
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.db.models import Count, Sum
from .models import Student, StaffProfile, TranscriptRequest, Payment


def superadmin_login(request):
    """SuperAdmin Login using superuser credentials"""
    if request.user.is_authenticated and request.user.is_superuser:
        return redirect('superadmin_dashboard')

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)

        if user is not None and user.is_superuser:
            login(request, user)
            messages.success(request, f"Welcome back, {user.username}!")
            return redirect('superadmin_dashboard')
        else:
            messages.error(request, "Invalid credentials or not a superuser.")
    
    return render(request, 'admin/superadmin_login.html')


from django.shortcuts import render
from django.contrib.auth.decorators import user_passes_test
from django.db.models import Sum
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from .models import Student, TranscriptRequest, TranscriptStatus, Transcript
from django.http import FileResponse, Http404
import os
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.barcode.qr import QrCodeWidget
from reportlab.graphics import renderPM
import urllib.parse


@user_passes_test(lambda u: u.is_superuser)
def superadmin_dashboard(request):
    """
    Super Admin Dashboard — shows overview, payments, transcript requests, and approved transcripts.
    """

    # ====== METRICS ======
    total_students = Student.objects.count()
    total_requests = TranscriptRequest.objects.count()
    total_approved = TranscriptStatus.objects.filter(stage="approved").count()
    total_pending = TranscriptStatus.objects.filter(stage="pending").count()

    total_revenue = (
        TranscriptRequest.objects.filter(payment_made=True)
        .aggregate(total=Sum('amount'))['total']
        or 0
    )

    # ====== RECENT PAYMENTS ======
    recent_payments = (
        TranscriptRequest.objects.filter(payment_made=True)
        .select_related('student')
        .order_by('-date_requested')[:20]
    )

    # ====== RECENT TRANSCRIPT REQUESTS ======
    requests_qs = (
        TranscriptRequest.objects.select_related('student')
        .prefetch_related('statuses', 'transcript')
        .order_by('-date_requested')[:20]
    )

    requests_data = []
    for req in requests_qs:
        latest_status = req.statuses.order_by('-updated_on', '-id').first()
        transcript_obj = getattr(req, 'transcript', None)
        # A transcript file may be stored in `file` (generated/merged) or `uploaded_file` (manual uploads).
        has_file = bool(transcript_obj and (getattr(transcript_obj, 'file', None) or getattr(transcript_obj, 'uploaded_file', None)))

        stage_map = {
            'pending': ("Awaiting Review", "info"),
            'accounts_office': ("At Accounts Office", "warning"),
            'exams_office': ("At Exams Office", "primary"),
            'registrar': ("At Registrar", "secondary"),
            'approved': ("Approved", "success"),
            'rejected': ("Rejected", "danger"),
        }
        stage_label = stage_map.get(latest_status.stage if latest_status else '', ("Processing", "dark"))

        # Prefer the merged/generated `file` field; fall back to `uploaded_file` when present.
        transcript_file_url = None
        if transcript_obj:
            try:
                if getattr(transcript_obj, 'file', None):
                    transcript_file_url = transcript_obj.file.url
                elif getattr(transcript_obj, 'uploaded_file', None):
                    transcript_file_url = transcript_obj.uploaded_file.url
            except Exception:
                transcript_file_url = None

        requests_data.append({
            'id': req.id,
            'student_name': req.student.name if req.student else '—',
            'index_number': req.student.index_number if req.student else '—',
            'transcript_type': req.transcript_type or '—',
            'date_requested': req.date_requested,
            'stage_label': stage_label[0],
            'stage_badge': stage_label[1],
            'download_ready': bool(transcript_file_url),
            'transcript_file': transcript_file_url,
            'transcript_id': getattr(transcript_obj, 'id', None),
        })

    # ====== APPROVED TRANSCRIPTS ======
    # Only include transcripts whose latest TranscriptStatus.stage is 'approved'.
    latest_status = TranscriptStatus.objects.filter(transcript_request=OuterRef('transcript_request')).order_by('-updated_on')
    approved_transcripts = (
        Transcript.objects.select_related('transcript_request__student')
        .annotate(latest_stage=Subquery(latest_status.values('stage')[:1]))
        .filter(latest_stage='approved')
        .filter(Q(file__isnull=False) | Q(uploaded_file__isnull=False))
        .distinct()
        .order_by('-date_generated')
    )

    context = {
        'total_students': total_students,
        'total_requests': total_requests,
        'total_approved': total_approved,
        'total_pending': total_pending,
        'total_revenue': total_revenue,
        'recent_payments': recent_payments,
        'requests_data': requests_data,
        'approved_transcripts': approved_transcripts,
    }

    return render(request, 'admin/superadmin_dashboard.html', context)


@user_passes_test(lambda u: u.is_superuser)
def superadmin_download_transcript(request, pk):
    """Allow superusers to securely download any transcript (generated or uploaded).

    Serves the file using FileResponse so that media storage permissions don't block access
    when using direct URLs.
    """
    transcript = get_object_or_404(Transcript, pk=pk)

    # Prefer generated/merged file, fall back to uploaded_file
    file_field = None
    if getattr(transcript, 'file', None):
        file_field = transcript.file
    elif getattr(transcript, 'uploaded_file', None):
        file_field = transcript.uploaded_file

    if not file_field:
        messages.error(request, "Transcript file not available for download.")
        return redirect('superadmin_dashboard')

    try:
        file_path = file_field.path
    except Exception:
        # Some storages (S3, etc.) may not provide a local path. Try to open via storage.
        try:
            fh = file_field.open('rb')
            response = HttpResponse(fh.read(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{transcript.transcript_request.reference_code}.pdf"'
            fh.close()
            return response
        except Exception:
            raise Http404("File not found")

    if not os.path.exists(file_path):
        raise Http404("File not found")

    filename = f"{transcript.transcript_request.reference_code}.pdf"
    response = FileResponse(open(file_path, 'rb'), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@user_passes_test(lambda u: u.is_superuser)
def superadmin_transcript_qr(request, pk):
    """Return a PNG QR code for the transcript verification URL for a given Transcript id.

    Uses ReportLab's QrCodeWidget and renderPM to produce a PNG image.
    """
    transcript = get_object_or_404(Transcript, pk=pk)
    tr_request = getattr(transcript, 'transcript_request', None)
    if not tr_request:
        raise Http404("Transcript request not found")

    # Build verification code used by verify view: index + cleaned reference
    index = getattr(getattr(tr_request, 'student', None), 'index_number', '')
    ref = tr_request.reference_code or ''
    ref_clean = urllib.parse.quote_plus(ref)

    base_url = request.build_absolute_uri('/')[:-1]
    code = f"{index}{ref_clean}"
    encoded = urllib.parse.quote(code, safe='')
    verify_url = f"{base_url}/verify/{encoded}/"

    try:
        qr = QrCodeWidget(verify_url)
        size = 200
        d = Drawing(size, size)
        d.add(qr)
        png = renderPM.drawToString(d, fmt='PNG')
        return HttpResponse(png, content_type='image/png')
    except Exception as e:
        # As a fallback, return 404
        raise Http404("QR generation failed")


# ====== EXPORT TO EXCEL ======
@user_passes_test(lambda u: u.is_superuser)
def export_payments_excel(request):
    """
    Exports all payments to an Excel file grouped by payment date.
    """
    payments = (
        TranscriptRequest.objects.filter(payment_made=True)
        .select_related('student')
        .order_by('-date_requested')
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Payments Report"

    headers = ["Date", "Student", "Index Number", "Transcript Type", "Amount (GHS)", "Reference Code"]
    ws.append(headers)

    current_day = None
    for pay in payments:
        date_str = pay.date_requested.strftime("%Y-%m-%d")
        if date_str != current_day:
            ws.append([f"--- {date_str} ---"])
            current_day = date_str

        ws.append([
            pay.date_requested.strftime("%Y-%m-%d %H:%M"),
            getattr(pay.student, 'name', '—'),
            getattr(pay.student, 'index_number', '—'),
            pay.transcript_type,
            float(pay.amount),
            pay.reference_code
        ])

    for col_num, _ in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 25

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = f'attachment; filename="Payments_Report_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response





def superadmin_logout(request):
    logout(request)
    messages.info(request, "You have been logged out.")
    return redirect('superadmin_login')



from django.contrib.auth.decorators import user_passes_test
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import FacultyRegistrar, TranscriptType, Department, Program
from .forms import FacultyRegistrarForm, TranscriptTypeForm, DepartmentForm, ProgramForm

def superadmin_required(view_func):
    return user_passes_test(lambda u: u.is_superuser, login_url='login')(view_func)


# app/views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .models import FacultyRegistrar
from .forms import FacultyRegistrarForm
from .decorators import superadmin_required  # your decorator

@superadmin_required
def manage_faculty_registrars(request):
    registrars = FacultyRegistrar.objects.all().order_by('faculty_name')

    # === CREATE ===
    if request.method == "POST" and 'create_registrar' in request.POST:
        form = FacultyRegistrarForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Faculty Registrar added successfully.")
            return redirect('manage_faculty_registrars')
    else:
        form = FacultyRegistrarForm()

    # === UPDATE ===
    if request.method == "POST" and 'update_registrar' in request.POST:
        registrar = get_object_or_404(FacultyRegistrar, id=request.POST.get('reg_id'))
        update_form = FacultyRegistrarForm(request.POST, request.FILES, instance=registrar)
        if update_form.is_valid():
            update_form.save()
            messages.success(request, "Faculty Registrar updated successfully.")
            return redirect('manage_faculty_registrars')

    # === DELETE ===
    if request.method == "POST" and 'delete_registrar' in request.POST:
        registrar = get_object_or_404(FacultyRegistrar, id=request.POST.get('reg_id'))
        registrar.delete()
        messages.success(request, "Faculty Registrar deleted successfully.")
        return redirect('manage_faculty_registrars')

    return render(request, 'admin/manage_faculty_registrars.html', {
        'form': form,
        'registrars': registrars,
    })


@superadmin_required
def manage_transcript_types(request):
    types = TranscriptType.objects.all().order_by('id')
    form = TranscriptTypeForm()

    # CREATE
    if request.method == "POST" and "create_type" in request.POST:
        form = TranscriptTypeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Transcript Type added successfully.")
            return redirect('manage_transcript_types')

    # UPDATE (only price allowed)
    if request.method == "POST" and "update_type" in request.POST:
        type_id = request.POST.get("type_id")
        transcript_type = get_object_or_404(TranscriptType, id=type_id)
        new_price = request.POST.get("price")

        # only update the price field
        if new_price:
            transcript_type.price = new_price
            transcript_type.save()
            messages.success(request, f"Price for '{transcript_type.get_type_display()}' updated successfully.")
        return redirect('manage_transcript_types')

    # DELETE
    if request.method == "POST" and "delete_type" in request.POST:
        type_id = request.POST.get("type_id")
        transcript_type = get_object_or_404(TranscriptType, id=type_id)
        transcript_type.delete()
        messages.success(request, "Transcript Type deleted successfully.")
        return redirect('manage_transcript_types')

    return render(request, 'admin/manage_transcript_types.html', {
        'form': form,
        'types': types
    })


@superadmin_required
def manage_departments(request):
    from django.shortcuts import get_object_or_404
    departments = Department.objects.all()
    form = DepartmentForm(request.POST or None)

    # CREATE
    if request.method == "POST" and "create_department" in request.POST:
        if form.is_valid():
            form.save()
            messages.success(request, "Department added successfully.")
            return redirect('manage_departments')

    # UPDATE
    if request.method == "POST" and "update_department" in request.POST:
        dep_id = request.POST.get("dep_id")
        department = get_object_or_404(Department, id=dep_id)
        form = DepartmentForm(request.POST, instance=department)
        if form.is_valid():
            form.save()
            messages.success(request, "Department updated successfully.")
            return redirect('manage_departments')

    # DELETE
    if request.method == "POST" and "delete_department" in request.POST:
        dep_id = request.POST.get("dep_id")
        department = get_object_or_404(Department, id=dep_id)
        department.delete()
        messages.success(request, "Department deleted successfully.")
        return redirect('manage_departments')

    return render(request, 'admin/manage_departments.html', {
        'form': form,
        'departments': departments,
    })


@superadmin_required
def manage_programs(request):
    programs = Program.objects.select_related('department').all()
    form = ProgramForm(request.POST or None)

    # CREATE
    if request.method == "POST" and "create_program" in request.POST and form.is_valid():
        form.save()
        messages.success(request, "Program added successfully.")
        return redirect('manage_programs')

    # UPDATE
    if request.method == "POST" and "update_program" in request.POST:
        prog_id = request.POST.get("prog_id")
        prog = Program.objects.get(id=prog_id)
        prog.name = request.POST.get("name")
        prog.department_id = request.POST.get("department")
        prog.save()
        messages.success(request, "Program updated successfully.")
        return redirect('manage_programs')

    # DELETE
    if request.method == "POST" and "delete_program" in request.POST:
        prog_id = request.POST.get("prog_id")
        Program.objects.filter(id=prog_id).delete()
        messages.success(request, "Program deleted successfully.")
        return redirect('manage_programs')

    return render(request, 'admin/manage_programs.html', {'form': form, 'programs': programs})


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from .models import WhatsAppGroup
from .forms import WhatsAppGroupForm

@staff_member_required
def manage_whatsapp_groups(request):
    groups = WhatsAppGroup.objects.all()
    form = WhatsAppGroupForm()

    # CREATE
    if request.method == 'POST' and 'create_group' in request.POST:
        form = WhatsAppGroupForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "WhatsApp group added successfully.")
            return redirect('manage_whatsapp_groups')

    # UPDATE
    if request.method == 'POST' and 'update_group' in request.POST:
        group_id = request.POST.get('group_id')
        group = get_object_or_404(WhatsAppGroup, id=group_id)
        form = WhatsAppGroupForm(request.POST, instance=group)
        if form.is_valid():
            form.save()
            messages.success(request, "Group updated successfully.")
            return redirect('manage_whatsapp_groups')

    # DELETE
    if request.method == 'POST' and 'delete_group' in request.POST:
        group_id = request.POST.get('group_id')
        group = get_object_or_404(WhatsAppGroup, id=group_id)
        group.delete()
        messages.success(request, "Group deleted successfully.")
        return redirect('manage_whatsapp_groups')

    context = {'form': form, 'groups': groups}
    return render(request, 'admin/manage_whatsapp_groups.html', context)


@superadmin_required
def manage_students(request):
    department = request.GET.get('department')
    program = request.GET.get('program')

    students = Student.objects.all().order_by('index_number')

    if department:
        students = students.filter(department__id=department)
    if program:
        students = students.filter(program__id=program)

    departments = Department.objects.all()
    programs = Program.objects.all()

    transcript_data = {}
    for student in students:
        transcript_requests = TranscriptRequest.objects.filter(
            student=student
        ).order_by('-date_requested')
        transcript_data[student.id] = transcript_requests

    return render(request, 'admin/manage_students.html', {
        'students': students,
        'transcript_data': transcript_data,
        'departments': departments,
        'programs': programs,
        'selected_department': department,
        'selected_program': program,
    })



from django.shortcuts import render, redirect
from .models import Contact
from .forms import ContactForm


def landing_page(request):
    contacts = Contact.objects.all()
    return render(request, "admin/landing.html", {"contacts": contacts})



from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .models import Contact
from .forms import ContactForm

def manage_contacts(request):
    contacts = Contact.objects.all().order_by('department')
    form = ContactForm()

    if request.method == 'POST' and 'create_contact' in request.POST:
        form = ContactForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Contact added successfully.")
            return redirect('add_contact')
    if request.method == 'POST' and 'update_contact' in request.POST:
        contact_id = request.POST.get('contact_id')
        contact = get_object_or_404(Contact, id=contact_id)
        form = ContactForm(request.POST, instance=contact)
        if form.is_valid():
            form.save()
            messages.success(request, "Contact updated successfully.")
            return redirect('add_contact')

    if request.method == 'POST' and 'delete_contact' in request.POST:
        contact_id = request.POST.get('contact_id')
        contact = get_object_or_404(Contact, id=contact_id)
        contact.delete()
        messages.success(request, "Contact deleted successfully.")
        return redirect('add_contact')

    context = {
        'form': form,
        'contacts': contacts,
    }
    return render(request, 'admin/manage_contacts.html', context)



import os
from django.conf import settings
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from .models import StaffProfile
from .forms import UserForm, StaffProfileForm

def admin_required(view_func):
    return user_passes_test(lambda u: u.is_staff or u.is_superuser)(view_func)

@login_required
@admin_required
def staff_profile_list(request):
    staff_profiles = StaffProfile.objects.select_related('user').all()
    return render(request, 'staff/staff_profile_list.html', {'staff_profiles': staff_profiles})


import os
from django.conf import settings
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from .models import StaffProfile
from .forms import UserForm, StaffProfileForm

def admin_required(view_func):
    """Only allow staff or superusers to edit profiles."""
    return user_passes_test(lambda u: u.is_staff or u.is_superuser)(view_func)

@login_required
@admin_required
def edit_staff_profile(request, staff_id):
    staff_profile = get_object_or_404(StaffProfile, staff_id=staff_id)
    user = staff_profile.user

    if request.method == 'POST':
        user_form = UserForm(request.POST, instance=user)
        profile_form = StaffProfileForm(request.POST, request.FILES, instance=staff_profile)

        if user_form.is_valid() and profile_form.is_valid():
            user_form.save()
            profile = profile_form.save()

            # If Vice Chancellor or Registrar, allow uploading a dedicated signature image
            if profile.role in ['vice_chancellor', 'registrar'] and 'signature' in request.FILES:
                try:
                    profile.signature = request.FILES['signature']
                    profile.save(update_fields=['signature'])
                    messages.success(request, f"The {profile.get_role_display()} signature has been updated.")
                except Exception as e:
                    messages.error(request, f"Failed to save signature: {e}")

            messages.success(request, "Staff profile updated successfully.")
            return redirect('staff_profile_list')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        user_form = UserForm(instance=user)
        profile_form = StaffProfileForm(instance=staff_profile)

    return render(request, 'staff/edit_staff_profile.html', {
        'user_form': user_form,
        'profile_form': profile_form,
        'staff_profile': staff_profile
    })


@login_required
@admin_required
def edit_officer(request, role):
    """Edit global officer record (vice_chancellor or registrar) and replace official static image."""
    role = role.lower()
    if role not in ('vice_chancellor', 'registrar'):
        messages.error(request, 'Invalid officer role.')
        return redirect('staff_profile_list')

    # Try to find the StaffProfile for that role; if missing, allow creating one
    staff_profile = StaffProfile.objects.filter(role=role).select_related('user').first()

    # If none exists, create a placeholder user + profile when POSTing; otherwise require an existing account
    if request.method == 'POST':
        # If profile exists, bind forms to it; else create new User via UserForm/StaffProfileForm logic
        if staff_profile:
            user = staff_profile.user
            user_form = UserForm(request.POST, instance=user)
            profile_form = StaffProfileForm(request.POST, request.FILES, instance=staff_profile)
        else:
            # Create a lightweight user from posted first_name/last_name/email fields
            from django.contrib.auth.models import User
            user_form = UserForm(request.POST)
            # Build a new StaffProfile instance but don't save yet
            profile_form = StaffProfileForm(request.POST, request.FILES)

        if user_form.is_valid() and profile_form.is_valid():
            user = user_form.save()
            profile = profile_form.save(commit=False)
            # If there was no existing profile, attach to the saved user
            if not staff_profile:
                profile.user = user
            profile.role = role
            profile.staff_id = profile.staff_id or (f"{role.upper()}-OFFICER")
            profile.save()

            # Handle uploaded signature file and store it on the profile
            if 'signature' in request.FILES:
                try:
                    profile.signature = request.FILES['signature']
                    profile.save(update_fields=['signature'])
                    messages.success(request, f"The {role.replace('_', ' ').title()} signature was updated.")
                except Exception as e:
                    messages.error(request, f"Failed to update signature: {e}")

            messages.success(request, f"{role.replace('_', ' ').title()} details saved.")
            return redirect('staff_profile_list')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        if staff_profile:
            user_form = UserForm(instance=staff_profile.user)
            profile_form = StaffProfileForm(instance=staff_profile)
        else:
            user_form = UserForm()
            profile_form = StaffProfileForm(initial={'role': role})

    display_name = role.replace('_', ' ').title()
    return render(request, 'staff/edit_officer.html', {
        'user_form': user_form,
        'profile_form': profile_form,
        'role': role,
        'staff_profile': staff_profile,
        'display_name': display_name,
    })


@login_required
@admin_required
def officers_list(request):
    """List the Vice Chancellor and Registrar entries with edit links."""
    officers = StaffProfile.objects.filter(role__in=['vice_chancellor', 'registrar']).select_related('user')
    # Ensure we include placeholders for roles that don't have profiles yet
    roles = ['vice_chancellor', 'registrar']
    role_map = {}
    for r in roles:
        profile = officers.filter(role=r).first()
        display = r.replace('_', ' ').title()
        role_map[r] = {
            'profile': profile,
            'display_name': display,
        }
    return render(request, 'staff/officers_list.html', {'role_map': role_map})


