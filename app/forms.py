from django import forms
from .models import (
    Student,
    Department,
    TranscriptRequest,
    Payment,
    TranscriptStatus,
    Transcript,
    TranscriptApproval,
    TranscriptBatch,
    TranscriptSelection,
    TranscriptReview,
    FacultyRegistrar,
)
import csv
from io import TextIOWrapper
import openpyxl

class TranscriptRequestForm(forms.ModelForm):
    class Meta:
        model = TranscriptRequest
        fields = ['student', 'transcript_type']
        widgets = {
            'student': forms.Select(attrs={'class': 'form-control'}),
            'transcript_type': forms.Select(attrs={'class': 'form-control'}),
        }

    def __init__(self, *args, **kwargs):
        user = kwargs.pop('user', None)
        super().__init__(*args, **kwargs)
        if user and not user.is_staff:
            self.fields.pop('student', None)

class PaymentForm(forms.ModelForm):
    class Meta:
        model = Payment
        fields = ['cleared', 'officer_name', 'notes']
        widgets = {
            'cleared': forms.CheckboxInput(attrs={'class': 'form-check-input'}),
            'officer_name': forms.TextInput(attrs={'class': 'form-control'}),
            'notes': forms.Textarea(attrs={'class': 'form-control', 'rows': 3}),
        }

from django import forms
from django.contrib.auth.models import User
from .models import StudentProfile, StaffProfile

class UserUpdateForm(forms.ModelForm):
    class Meta:
        model = User
        fields = ['email']
        widgets = {
            'email': forms.EmailInput(attrs={'class': 'form-control'}),
        }

class StudentProfileForm(forms.ModelForm):
    class Meta:
        model = StudentProfile
        fields = ['phone_number', 'profile_image']
        widgets = {
            'phone_number': forms.TextInput(attrs={'class': 'form-control'}),
            'profile_image': forms.ClearableFileInput(attrs={'class': 'form-control'}),
        }

class StaffProfileForm(forms.ModelForm):
    class Meta:
        model = StaffProfile
        fields = ['phone_number', 'profile_image']
        widgets = {
            'phone_number': forms.TextInput(attrs={'class': 'form-control'}),
            'profile_image': forms.ClearableFileInput(attrs={'class': 'form-control'}),
        }


from django import forms
from .models import FacultyRegistrar

class FacultyRegistrarForm(forms.ModelForm):
    class Meta:
        model = FacultyRegistrar
        fields = ['name', 'signature', 'faculty_name']
        widgets = {
            'name': forms.TextInput(attrs={
                'class': 'form-control',
                'placeholder': 'Enter registrar’s full name'
            }),
            'faculty_name': forms.TextInput(attrs={
                'class': 'form-control',
                'placeholder': 'Enter faculty name (e.g., Faculty of Engineering)'
            }),
            'signature': forms.ClearableFileInput(attrs={
                'class': 'form-control'
            }),
        }



class TranscriptStatusForm(forms.ModelForm):
    class Meta:
        model = TranscriptStatus
        fields = ['stage', 'remarks', 'updated_by']
        widgets = {
            'stage': forms.Select(attrs={'class': 'form-control'}),
            'remarks': forms.Textarea(attrs={'class': 'form-control', 'rows': 3}),
            'updated_by': forms.TextInput(attrs={'class': 'form-control'}),
        }





class TranscriptForm(forms.ModelForm):
    class Meta:
        model = Transcript
        fields = ['file', 'generated_by', 'dean_signature', 'registrar_signature']
        widgets = {
            'file': forms.FileInput(attrs={'class': 'form-control'}),
            'generated_by': forms.TextInput(attrs={'class': 'form-control'}),
        }


from django import forms
from .models import TranscriptApproval

class TranscriptApprovalForm(forms.ModelForm):
    class Meta:
        model = TranscriptApproval
        fields = ['approved', 'remarks']  # removed 'approved_by' from here
        widgets = {
            'approved': forms.CheckboxInput(attrs={'class': 'form-check-input'}),
            'remarks': forms.Textarea(attrs={'class': 'form-control', 'rows': 3}),
        }

from .models import FeeClearance

class FeeClearanceForm(forms.ModelForm):
    class Meta:
        model = FeeClearance
        fields = ['cleared', 'owes', 'amount_owed', 'remarks', 'invoice_file']
        widgets = {
            'cleared': forms.CheckboxInput(attrs={'class': 'form-check-input'}),
            'owes': forms.CheckboxInput(attrs={'class': 'form-check-input'}),
            'amount_owed': forms.NumberInput(attrs={'class': 'form-control', 'step': '0.01', 'min': '0'}),
            'remarks': forms.Textarea(attrs={'class': 'form-control', 'rows': 2, 'placeholder': 'Enter remarks...'}),
            'invoice_file': forms.ClearableFileInput(attrs={'class': 'form-control'}),
        }


from django import forms
from .models import Student, Department
import csv
from io import TextIOWrapper
import openpyxl

class BulkStudentUploadForm(forms.Form):
    file = forms.FileField(
        help_text="Upload a CSV or Excel (.xlsx) file with columns: name, index_number, date_entered, date_completed, department, owes_fees"
    )

    def process_file(self):
        uploaded_file = self.cleaned_data['file']
        filename = uploaded_file.name.lower()

        students = []

        if filename.endswith('.csv'):
            file = TextIOWrapper(uploaded_file.file, encoding='utf-8')
            reader = csv.DictReader(file)
            for row in reader:
                try:
                    department = Department.objects.get(department=row.get('department'))
                    students.append(Student(
                        name=row.get('name'),
                        index_number=row.get('index_number'),
                        date_entered=row.get('date_entered'),
                        date_completed=row.get('date_completed'),
                        department=department,
                        owes_fees=row.get('owes_fees', '').lower() in ['true', '1', 'yes'],
                    ))
                except Department.DoesNotExist:
                    continue

        elif filename.endswith('.xlsx'):
            wb = openpyxl.load_workbook(uploaded_file)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                data = dict(zip(headers, row))
                try:
                    department = Department.objects.get(department=data.get('department'))
                    students.append(Student(
                        name=data.get('name'),
                        index_number=data.get('index_number'),
                        date_entered=data.get('date_entered'),
                        date_completed=data.get('date_completed'),
                        department=department,
                        owes_fees=str(data.get('owes_fees')).lower() in ['true', '1', 'yes'],
                    ))
                except Department.DoesNotExist:
                    continue

        return students



class StudentForm(forms.ModelForm):
    class Meta:
        model = Student
        fields = ['name', 'index_number', 'date_entered', 'date_completed', 'department', 'owes_fees']
        widgets = {
            'name': forms.TextInput(attrs={'class': 'form-control'}),
            'index_number': forms.TextInput(attrs={'class': 'form-control'}),
            'date_entered': forms.DateInput(attrs={'class': 'form-control', 'type': 'date'}),
            'date_completed': forms.DateInput(attrs={'class': 'form-control', 'type': 'date'}),
            'department': forms.Select(attrs={'class': 'form-control'}),
            'owes_fees': forms.CheckboxInput(attrs={'class': 'form-check-input'}),
        }



from django import forms
from django.contrib.auth.models import User
from .models import StudentProfile
import re

from django.contrib.auth.forms import PasswordResetForm, SetPasswordForm


class ResetPasswordForm(PasswordResetForm):
    """Wrapper around Django's PasswordResetForm to provide bootstrap widgets and keep project forms colocated."""
    email = forms.EmailField(widget=forms.EmailInput(attrs={'class': 'form-control', 'placeholder': 'you@example.com'}))


class ResetPasswordConfirmForm(SetPasswordForm):
    """Wrapper around Django's SetPasswordForm to style widgets and keep consistent labels."""
    new_password1 = forms.CharField(widget=forms.PasswordInput(attrs={'class': 'form-control'}), label='New password')
    new_password2 = forms.CharField(widget=forms.PasswordInput(attrs={'class': 'form-control'}), label='Confirm new password')

class CustomSignupForm(forms.ModelForm):
    password1 = forms.CharField(
        widget=forms.PasswordInput(attrs={'class': 'form-control'}),
        label="Password"
    )
    password2 = forms.CharField(
        widget=forms.PasswordInput(attrs={'class': 'form-control'}),
        label="Confirm Password"
    )
    index_number = forms.CharField(
        widget=forms.TextInput(attrs={'class': 'form-control'}),
        label="Index Number"
    )
    phone = forms.CharField(
        widget=forms.TextInput(attrs={'class': 'form-control', 'placeholder': '233XXXXXXXXX'}),
        label="Phone Number"
    )

    class Meta:
        model = User
        fields = ['username', 'email']
        widgets = {
            'username': forms.TextInput(attrs={'class': 'form-control'}),
            'email': forms.EmailInput(attrs={'class': 'form-control'}),
        }

    def clean_phone(self):
        phone = self.cleaned_data.get('phone')
        if not re.match(r'^(?:\+233|233|0)\d{9}$', phone):
            raise forms.ValidationError("Enter a valid Ghanaian phone number.")
        if phone.startswith('0'):
            phone = '233' + phone[1:]
        elif phone.startswith('+233'):
            phone = phone[1:]
        return phone

    def clean(self):
        cleaned_data = super().clean()
        if cleaned_data.get('password1') != cleaned_data.get('password2'):
            raise forms.ValidationError("Passwords do not match.")
        return cleaned_data

    def save(self, commit=True):
        user = super().save(commit=False)
        user.set_password(self.cleaned_data['password1'])
        if commit:
            user.save()
            StudentProfile.objects.create(
                user=user,
                index_number=self.cleaned_data['index_number'],
                phone_number=self.cleaned_data['phone']
            )
        return user

    def clean_password2(self):
        password = self.cleaned_data.get("password1")
        confirm_password = self.cleaned_data.get("password2")

        if password != confirm_password:
            raise forms.ValidationError("Passwords do not match.")

        if len(password) < 8:
            raise forms.ValidationError("Password must be at least 8 characters long.")
        if not re.search(r"[A-Z]", password):
            raise forms.ValidationError("Password must contain at least one uppercase letter.")
        if not re.search(r"[a-z]", password):
            raise forms.ValidationError("Password must contain at least one lowercase letter.")
        if not re.search(r"[0-9]", password):
            raise forms.ValidationError("Password must contain at least one number.")
        if not re.search(r"[!@#$%^&*(),.?\":{}|<>]", password):
            raise forms.ValidationError("Password must contain at least one special character.")

        return confirm_password

    def clean_index_number(self):
        index = self.cleaned_data.get("index_number")
        if StudentProfile.objects.filter(index_number=index).exists():
            raise forms.ValidationError("This index number is already registered.")
        return index

    def clean_phone(self):
        phone = self.cleaned_data.get("phone")
        if not phone.isdigit():
            raise forms.ValidationError("Phone number must contain only digits.")
        if StudentProfile.objects.filter(phone_number=phone).exists():
            raise forms.ValidationError("This phone number is already registered.")
        return phone

    def save(self, commit=True):
        user = super().save(commit=False)
        user.set_password(self.cleaned_data["password1"])
        user.is_staff = False
        user.is_superuser = False

        if commit:
            user.save()
            index_number = self.cleaned_data["index_number"]
            phone = self.cleaned_data["phone"]

            student_instance = Student.objects.filter(index_number=index_number).first()

            StudentProfile.objects.create(
                user=user,
                student=student_instance,
                index_number=index_number,
                phone_number=phone
            )

        return user


from django import forms
from django.contrib.auth.models import User
from .models import StaffProfile
import re

class StaffSignupForm(forms.ModelForm):
    first_name = forms.CharField(widget=forms.TextInput(attrs={'class': 'form-control'}))
    last_name = forms.CharField(widget=forms.TextInput(attrs={'class': 'form-control'}))
    password1 = forms.CharField(widget=forms.PasswordInput(attrs={'class': 'form-control'}), label="Password")
    password2 = forms.CharField(widget=forms.PasswordInput(attrs={'class': 'form-control'}), label="Confirm Password")
    staff_id = forms.CharField(widget=forms.TextInput(attrs={'class': 'form-control'}))
    role = forms.ChoiceField(choices=StaffProfile.ROLE_CHOICES, widget=forms.Select(attrs={'class': 'form-select'}))

    class Meta:
        model = User
        fields = []

    def clean_staff_id(self):
        staff_id = self.cleaned_data.get("staff_id")
        if User.objects.filter(username=staff_id).exists():
            raise forms.ValidationError("A staff member with this ID already exists.")
        return staff_id

    def clean_password2(self):
        password = self.cleaned_data.get("password1")
        confirm = self.cleaned_data.get("password2")
        if password != confirm:
            raise forms.ValidationError("Passwords do not match.")
        if len(password) < 8:
            raise forms.ValidationError("Password must be at least 8 characters.")
        if not re.search(r"[A-Z]", password):
            raise forms.ValidationError("Password must contain at least one uppercase letter.")
        if not re.search(r"[a-z]", password):
            raise forms.ValidationError("Password must contain at least one lowercase letter.")
        if not re.search(r"[0-9]", password):
            raise forms.ValidationError("Password must contain at least one number.")
        if not re.search(r"[!@#$%^&*(),.?\":{}|<>]", password):
            raise forms.ValidationError("Password must contain at least one special character.")
        return confirm

    def save(self, commit=True):
        user = User(
            username=self.cleaned_data['staff_id'],
            first_name=self.cleaned_data['first_name'],
            last_name=self.cleaned_data['last_name'],
            is_staff=True
        )
        user.set_password(self.cleaned_data['password1'])
        if commit:
            user.save()
            StaffProfile.objects.create(
                user=user,
                staff_id=self.cleaned_data['staff_id'],
                role=self.cleaned_data['role']
            )
        return user

    
    
    from django import forms

class StudentUploadForm(forms.Form):
    file = forms.FileField(
        label="Select CSV or Excel File",
        widget=forms.ClearableFileInput(attrs={'class': 'form-control'})
    )
    
    
from .models import Student



class StudentUploadForm(forms.Form):
    file = forms.FileField(label="Select CSV or Excel file")

class StudentForm(forms.ModelForm):
    class Meta:
        model = Student
        fields = ['name', 'index_number', 'date_entered', 'date_completed', 'program', 'owes_fees']
        widgets = {
            'date_entered': forms.DateInput(attrs={'type': 'date', 'class': 'form-control'}),
            'date_completed': forms.DateInput(attrs={'type': 'date', 'class': 'form-control'}),
            'name': forms.TextInput(attrs={'class': 'form-control'}),
            'index_number': forms.TextInput(attrs={'class': 'form-control'}),
            'program': forms.Select(attrs={'class': 'form-control'}),
            'owes_fees': forms.CheckboxInput(attrs={'class': 'form-check-input'}),
        }


class TranscriptBatchForm(forms.ModelForm):
    class Meta:
        model = TranscriptBatch
        fields = ['notes']
        widgets = {
            'notes': forms.Textarea(attrs={'class': 'form-control', 'rows': 3, 'placeholder': 'Add notes for this batch...'}),
        }


class PasswordResetRequestForm(forms.Form):
    email = forms.EmailField(widget=forms.EmailInput(attrs={'class': 'form-control'}))


class PasswordResetConfirmForm(forms.Form):
    code = forms.CharField(widget=forms.TextInput(attrs={'class': 'form-control'}))
    # Accept either the email or the student's index number to bind the code to the user
    email_or_index = forms.CharField(
        required=False,
        widget=forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'Email or index (optional)'})
    )
    new_password1 = forms.CharField(widget=forms.PasswordInput(attrs={'class': 'form-control'}), label='New password')
    new_password2 = forms.CharField(widget=forms.PasswordInput(attrs={'class': 'form-control'}), label='Confirm new password')

    def clean(self):
        cleaned = super().clean()
        p1 = cleaned.get('new_password1')
        p2 = cleaned.get('new_password2')
        if p1 and p2 and p1 != p2:
            raise forms.ValidationError('Passwords do not match.')
        return cleaned


class TranscriptSelectionForm(forms.Form):
    """Form for selecting transcripts to include in a batch"""
    transcript_requests = forms.ModelMultipleChoiceField(
        queryset=TranscriptRequest.objects.none(),
        widget=forms.CheckboxSelectMultiple(attrs={'class': 'form-check-input'}),
        required=True
    )
    
    def __init__(self, *args, **kwargs):
        available_requests = kwargs.pop('available_requests', None)
        super().__init__(*args, **kwargs)
        if available_requests:
            self.fields['transcript_requests'].queryset = available_requests


class TranscriptReviewForm(forms.ModelForm):
    class Meta:
        model = TranscriptReview
        fields = ['approved', 'comments', 'changes_made']
        widgets = {
            'approved': forms.CheckboxInput(attrs={'class': 'form-check-input'}),
            'comments': forms.Textarea(attrs={'class': 'form-control', 'rows': 3, 'placeholder': 'Add review comments...'}),
            'changes_made': forms.Textarea(attrs={'class': 'form-control', 'rows': 3, 'placeholder': 'Describe any changes made...'}),
        }


class FacultyChangeForm(forms.Form):
    """Form for registrar to change faculty for unofficial transcripts"""
    faculty_registrar = forms.ModelChoiceField(
        queryset=FacultyRegistrar.objects.all(),
        empty_label="-- Select Faculty Registrar --",
        widget=forms.Select(attrs={'class': 'form-control'}),
        label="Faculty Registrar"
    )
    reason = forms.CharField(
        widget=forms.Textarea(attrs={'class': 'form-control', 'rows': 3, 'placeholder': 'Reason for changing faculty...'}),
        label="Reason for Change",
        required=True
    )


from django import forms
from .models import FacultyRegistrar, TranscriptType, Department, Program

class FacultyRegistrarForm(forms.ModelForm):
    class Meta:
        model = FacultyRegistrar
        fields = ['name', 'signature', 'faculty_name']


class TranscriptTypeForm(forms.ModelForm):
    class Meta:
        model = TranscriptType
        fields = ['type', 'price']


class DepartmentForm(forms.ModelForm):
    class Meta:
        model = Department
        fields = ['department', 'HoD']


class ProgramForm(forms.ModelForm):
    class Meta:
        model = Program
        fields = ['name', 'department']



from .models import WhatsAppGroup, Contact
class WhatsAppGroupForm(forms.ModelForm):
    class Meta:
        model = WhatsAppGroup
        fields = ['name', 'description', 'link', 'icon', 'active']
        widgets = {
            'name': forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'Enter group name'}),
            'description': forms.Textarea(attrs={'class': 'form-control', 'rows': 2, 'placeholder': 'Short description'}),
            'link': forms.URLInput(attrs={'class': 'form-control', 'placeholder': 'https://chat.whatsapp.com/...'}),
            'icon': forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'e.g. fa-users'}),
            'active': forms.CheckboxInput(attrs={'class': 'form-check-input'}),
        }


class ContactForm(forms.ModelForm):
    class Meta:
        model = Contact
        fields = ['department', 'phone_number', 'active']
        widgets = {
            'department': forms.TextInput(attrs={'class': 'form-control rounded-3', 'placeholder': 'Department'}),
            'phone_number': forms.TextInput(attrs={'class': 'form-control rounded-3', 'placeholder': 'Phone Number'}),
            'active': forms.CheckboxInput(attrs={'class': 'form-check-input'}),
        }

from django import forms
from django.contrib.auth.models import User
from .models import StaffProfile

class UserForm(forms.ModelForm):
    class Meta:
        model = User
        fields = ['first_name', 'last_name', 'email']
        widgets = {
            'first_name': forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'First Name'}),
            'last_name': forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'Last Name'}),
            'email': forms.EmailInput(attrs={'class': 'form-control', 'placeholder': 'Email'}),
        }

class StaffProfileForm(forms.ModelForm):
    class Meta:
        model = StaffProfile
        fields = ['staff_id', 'role', 'phone_number', 'profile_image', 'signature']
        widgets = {
            'staff_id': forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'Staff ID'}),
            'role': forms.Select(attrs={'class': 'form-select'}),
            'phone_number': forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'Phone Number'}),
            'profile_image': forms.FileInput(attrs={'class': 'form-control'}),
            'signature': forms.FileInput(attrs={'class': 'form-control'}),
        }


class RegistrarUploadForm(forms.Form):
    student_index = forms.CharField(required=True, widget=forms.TextInput(attrs={'class': 'form-control'}))
    transcript_file = forms.FileField(required=True, widget=forms.ClearableFileInput(attrs={'class': 'form-control'}))
    transcript_type = forms.ChoiceField(choices=[('official','Official'), ('unofficial','Unofficial')], widget=forms.Select(attrs={'class':'form-control'}))
    faculty_registrar = forms.ModelChoiceField(queryset=FacultyRegistrar.objects.all(), required=False, widget=forms.Select(attrs={'class':'form-control'}))
    amount_received = forms.DecimalField(required=False, max_digits=10, decimal_places=2, widget=forms.NumberInput(attrs={'class':'form-control'}))
    amount_accounts_office = forms.DecimalField(required=False, max_digits=10, decimal_places=2, widget=forms.NumberInput(attrs={'class':'form-control'}))
    amount_superadmin = forms.DecimalField(required=False, max_digits=10, decimal_places=2, widget=forms.NumberInput(attrs={'class':'form-control'}))
    amount_registrar = forms.DecimalField(required=False, max_digits=10, decimal_places=2, widget=forms.NumberInput(attrs={'class':'form-control'}))


class ExamsOfficeUploadForm(forms.Form):
    """Form for exams office manual transcript processing"""
    student_index = forms.CharField(required=True, widget=forms.TextInput(attrs={'class': 'form-control'}))
    transcript_file = forms.FileField(required=True, widget=forms.ClearableFileInput(attrs={'class': 'form-control'}))
    transcript_type = forms.ChoiceField(choices=[('official','Official'), ('unofficial','Unofficial')], widget=forms.Select(attrs={'class':'form-control'}))
    faculty_registrar = forms.ModelChoiceField(queryset=FacultyRegistrar.objects.all(), required=False, widget=forms.Select(attrs={'class':'form-control'}))
    remarks = forms.CharField(required=False, widget=forms.Textarea(attrs={'class': 'form-control', 'rows': 3, 'placeholder': 'Add remarks for registrar review...'}))





